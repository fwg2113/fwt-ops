#!/usr/bin/env python3
"""
FWT Customer & Appointment Import Script v2
=============================================
Yearly sheets are the source of truth. Google Calendar used ONLY for phone enrichment on 2018-2019.
"""

import openpyxl
import json
import re
import uuid
import sys
from datetime import datetime, date
from collections import defaultdict

XLSX_PATH = 'FWT Customers & Appointments/FWT Customer _ Appointment Database.xlsx'

# ============================================================================
# HELPERS
# ============================================================================
def normalize_phone(raw):
    if raw is None: return None
    s = str(raw).split('.')[0].strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) == 11 and digits.startswith('1'): digits = digits[1:]
    return digits if len(digits) == 10 else None

def normalize_email(raw):
    if raw is None: return None
    s = str(raw).strip().lower()
    return s if '@' in s and '.' in s else None

def parse_name(raw):
    if not raw: return (None, None)
    name = str(raw).strip()
    if re.match(r'^\d', name): return (None, None)
    name = re.sub(r'^(mr\.?|mrs\.?|ms\.?|dr\.?)\s+', '', name, flags=re.IGNORECASE)
    if '/' in name: name = name.split('/')[0].strip()
    parts = [p for p in name.split() if not re.match(r'^\d{4}-\d{2}-\d{2}$', p)]
    if len(parts) == 0: return (None, None)
    if len(parts) == 1: return (parts[0].title(), None)
    return (parts[0].title(), ' '.join(parts[1:]).title())

def parse_vehicle_string(s):
    if not s: return (None, None, None)
    s = str(s).strip()
    if s.lower() in ('none', 'n/a', ''): return (None, None, None)
    m = re.match(r'(\d{4})\s+(.+)', s)
    if m:
        year = int(m.group(1))
        rest = m.group(2).strip().split()
        return (year, rest[0] if rest else None, ' '.join(rest[1:]) if len(rest) > 1 else None)
    parts = s.split()
    return (None, parts[0] if parts else None, ' '.join(parts[1:]) if len(parts) > 1 else None)

def parse_vehicle_from_deal(deal):
    if not deal: return (None, None, None)
    s = str(deal).strip()
    stop = r'(?:S\d|G\d|i3|BC|BLK|UP|UXP|PPF|TS|WS|WFI|PT|FULL|R&R|Removal|GC|MSI|WAITING|\$)'
    m = re.match(r'(\d{4})\s+(.+?)(?:\s+' + stop + r')', s, re.IGNORECASE)
    if m:
        year = int(m.group(1))
        parts = m.group(2).strip().split()
        return (year, parts[0] if parts else None, ' '.join(parts[1:]) if len(parts) > 1 else None)
    m2 = re.match(r'(\d{4})\s+(.+?)(?:\s+\$|$)', s)
    if m2:
        year = int(m2.group(1))
        parts = m2.group(2).strip().split()
        return (year, parts[0] if parts else None, ' '.join(parts[1:]) if len(parts) > 1 else None)
    return (None, None, None)

def safe_date(val):
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, date): return val.strftime('%Y-%m-%d')
    return None

def safe_float(val):
    try: return float(val) if val else 0.0
    except: return 0.0

def safe_str(val):
    if val is None: return None
    s = str(val).strip()
    return s if s and s.lower() not in ('none', 'false', 'true') else None

def esc(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def is_valid_name(n):
    if not n: return False
    if re.match(r'^\d', n): return False
    if re.match(r'^[\d\s./-]+$', n): return False
    return True

# ============================================================================
# DATA STRUCTURES
# ============================================================================
class Customer:
    def __init__(self, phone):
        self.phone = phone
        self.emails = []
        self.first_name = None
        self.last_name = None
        self.company = None
        self.appointments = []

    def add_email(self, email, dt=None):
        if email: self.emails.append((email, dt))

    def best_email(self):
        if not self.emails: return None
        return sorted(self.emails, key=lambda x: x[1] or '0000', reverse=True)[0][0]

    def update_name(self, first, last):
        if is_valid_name(first) and (not self.first_name or len(first) > len(self.first_name)):
            self.first_name = first
        if is_valid_name(last) and (not self.last_name or (len(last) > len(self.last_name))):
            self.last_name = last

class Appointment:
    def __init__(self):
        self.date = None
        self.vehicle_year = None
        self.vehicle_make = None
        self.vehicle_model = None
        self.appointment_desc = None  # Full appointment text from spreadsheet
        self.shade = None
        self.windshield = None
        self.tint_strip = None
        self.sun_roof = None
        self.removal = None
        self.wfi = None
        self.price = 0.0
        self.tip = 0.0
        self.total = 0.0
        self.payment = None
        self.source = None
        self.note = None
        self.company = None
        self.gc = None
        self.invoice_num = None
        self.event_id = None

    def dedup_key(self):
        return f"{self.date}|{self.vehicle_year}|{(self.vehicle_make or '').lower()}|{round(self.total)}"

# ============================================================================
# MAIN
# ============================================================================
def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    customers = {}  # phone -> Customer

    def get_or_create(phone):
        if phone not in customers:
            customers[phone] = Customer(phone)
        return customers[phone]

    # ========================================================================
    # PASS 0: Enrich emails from Google Calendar (no appointment data used)
    # ========================================================================
    print("Reading Google Calendar for email enrichment...")
    ws = wb['Google Calendar Data']
    gcal_enriched = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (list(row) + [None]*14)[:14]
        first_name, last_name, phone_raw, email_raw, visit_date = vals[:5]
        phone = normalize_phone(phone_raw)
        if not phone: continue
        email = normalize_email(email_raw)
        dt = safe_date(visit_date)
        first, _ = parse_name(first_name)
        last_parsed = str(last_name).strip().title() if last_name else None
        if email:
            c = get_or_create(phone)
            c.add_email(email, dt)
            if first: c.update_name(first, last_parsed)
            gcal_enriched += 1
    print(f"  -> {gcal_enriched} email enrichments")

    # ========================================================================
    # PASS 1: 2025 Appointments (cleanest data)
    # ========================================================================
    print("Reading 2025 Appointments...")
    ws = wb['2025 Appointments']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (list(row) + [None]*25)[:25]
        hs, dt_raw, name, phone_raw, email_raw = vals[0:5]
        appt_desc, vehicle, shade, ws_val, ts_val, sr_val, removal = vals[5:12]
        price, wfi, tip, total, payment, source, note, company, gc, invoice_num, record_id, event_id = vals[12:24]

        phone = normalize_phone(phone_raw)
        if not phone: continue
        dt = safe_date(dt_raw)
        if not dt: continue

        email = normalize_email(email_raw)
        first, last = parse_name(name)
        vy, vm, vmod = parse_vehicle_string(vehicle)

        c = get_or_create(phone)
        c.update_name(first, last)
        c.add_email(email, dt)
        if safe_str(company): c.company = safe_str(company)

        a = Appointment()
        a.date = dt
        a.vehicle_year = vy
        a.vehicle_make = vm
        a.vehicle_model = vmod
        a.appointment_desc = safe_str(appt_desc)
        a.shade = safe_str(shade)
        a.windshield = safe_str(ws_val)
        a.tint_strip = safe_str(ts_val)
        a.sun_roof = safe_str(sr_val)
        a.removal = safe_str(removal)
        a.price = safe_float(price)
        a.wfi = safe_float(wfi)
        a.tip = safe_float(tip)
        a.total = safe_float(total)
        a.payment = safe_str(payment)
        a.source = safe_str(source)
        a.note = safe_str(note)
        a.gc = safe_str(gc)
        a.invoice_num = safe_str(invoice_num)
        a.event_id = safe_str(event_id)
        c.appointments.append(a)
        count += 1
    print(f"  -> {count} appointments")

    # ========================================================================
    # PASS 2: 2024 Appointments (same format as 2025)
    # ========================================================================
    print("Reading 2024 Appointments...")
    ws = wb['2024 Appointments']
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip header + spacer row
        vals = (list(row) + [None]*24)[:24]
        hs, dt_raw, name, phone_raw, email_raw = vals[0:5]
        appt_desc, vehicle, shade, ws_val, ts_val, sr_val, col11 = vals[5:12]
        price, wfi, tip, total, payment, source, note, company, gc, invoice_num, record_id, event_id = vals[12:24]

        phone = normalize_phone(phone_raw)
        if not phone: continue
        dt = safe_date(dt_raw)
        if not dt: continue

        email = normalize_email(email_raw)
        first, last = parse_name(name)
        vy, vm, vmod = parse_vehicle_string(vehicle)

        c = get_or_create(phone)
        c.update_name(first, last)
        c.add_email(email, dt)
        if safe_str(company): c.company = safe_str(company)

        a = Appointment()
        a.date = dt
        a.vehicle_year = vy
        a.vehicle_make = vm
        a.vehicle_model = vmod
        a.appointment_desc = safe_str(appt_desc)
        a.shade = safe_str(shade)
        a.windshield = safe_str(ws_val)
        a.tint_strip = safe_str(ts_val)
        a.sun_roof = safe_str(sr_val)
        a.price = safe_float(price)
        a.wfi = safe_float(wfi)
        a.tip = safe_float(tip)
        a.total = safe_float(total)
        a.payment = safe_str(payment)
        a.source = safe_str(source)
        a.note = safe_str(note)
        a.gc = safe_str(gc)
        a.invoice_num = safe_str(invoice_num)
        a.event_id = safe_str(event_id)
        c.appointments.append(a)
        count += 1
    print(f"  -> {count} appointments")

    # ========================================================================
    # PASS 3: 2020-2023 HubSpot format
    # ========================================================================
    for year in ['2020', '2021', '2022', '2023']:
        print(f"Reading {year} Appointments...")
        ws = wb[f'{year} Appointments']
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = (list(row) + [None]*18)[:18]
            record_id, email_raw, first_name, last_name, phone_raw = vals[0:5]
            deal_name, film, close_date, total_price, gift_cert = vals[5:10]
            product_type, pipeline, source, dup_count, deal_stage, order_id = vals[10:16]

            phone = normalize_phone(phone_raw)
            if not phone: continue
            dt = safe_date(close_date)
            if not dt: continue

            email = normalize_email(email_raw)
            first, _ = parse_name(first_name)
            last = str(last_name).strip().title() if last_name and is_valid_name(str(last_name).strip()) else None

            vy, vm, vmod = parse_vehicle_from_deal(deal_name)

            c = get_or_create(phone)
            c.update_name(first, last)
            c.add_email(email, dt)

            a = Appointment()
            a.date = dt
            a.vehicle_year = vy
            a.vehicle_make = vm
            a.vehicle_model = vmod
            a.appointment_desc = safe_str(deal_name)
            a.shade = safe_str(film)
            a.total = safe_float(total_price)
            a.price = a.total
            a.source = safe_str(source)
            a.gc = 'Yes' if gift_cert else None
            c.appointments.append(a)
            count += 1
        print(f"  -> {count} appointments")

    # ========================================================================
    # PASS 4: HubSpot email list (email enrichment only, skip dupes)
    # ========================================================================
    print("Reading HubSpot email list for enrichment...")
    ws = wb['Hubspot_FWT_Email_List_with_Dup']
    enriched = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (list(row) + [None]*17)[:17]
        _, email_raw, first_name, last_name, phone_raw = vals[0:5]
        phone = normalize_phone(phone_raw)
        email = normalize_email(email_raw)
        if phone and email:
            c = get_or_create(phone)
            c.add_email(email, None)
            first, _ = parse_name(first_name)
            last = str(last_name).strip().title() if last_name and is_valid_name(str(last_name).strip()) else None
            c.update_name(first, last)
            enriched += 1
    print(f"  -> {enriched} customers enriched with emails")

    # 2018 and 2019 skipped -- no phone numbers, name-only matching is unreliable
    print("Skipping 2018 and 2019 (no phone numbers on those sheets)")

    wb.close()

    # ========================================================================
    # DEDUP per customer
    # ========================================================================
    print("\nDeduplicating...")
    total_before = sum(len(c.appointments) for c in customers.values())
    for phone, c in customers.items():
        seen = {}
        unique = []
        for a in c.appointments:
            key = a.dedup_key()
            if key not in seen:
                seen[key] = a
                unique.append(a)
            else:
                # Keep the one with more data (prefer 2024/2025 sheets over HubSpot)
                existing = seen[key]
                if a.appointment_desc and not existing.appointment_desc:
                    existing.appointment_desc = a.appointment_desc
                if a.shade and not existing.shade:
                    existing.shade = a.shade
                if a.payment and not existing.payment:
                    existing.payment = a.payment
                if a.event_id and not existing.event_id:
                    existing.event_id = a.event_id
                if a.note and not existing.note:
                    existing.note = a.note
        c.appointments = unique
    total_after = sum(len(c.appointments) for c in customers.values())
    print(f"  -> {total_before} -> {total_after} ({total_before - total_after} dupes removed)")

    # ========================================================================
    # GENERATE SQL
    # ========================================================================
    print(f"\nGenerating SQL for {len(customers)} customers, {total_after} appointments...")

    cust_sql = []
    book_sql = []
    seq_sql = []
    date_seqs = defaultdict(int)

    for phone, c in customers.items():
        cid = str(uuid.uuid4())
        email = c.best_email()
        first = c.first_name or ''
        last = c.last_name or ''

        dates = [a.date for a in c.appointments if a.date]
        totals = [a.total for a in c.appointments if a.total > 0]
        first_visit = min(dates) if dates else None
        last_visit = max(dates) if dates else None
        lifetime = sum(totals)

        cust_sql.append(
            f"INSERT INTO customers (id, phone, email, first_name, last_name, company_name, "
            f"lifetime_spend, visit_count, first_visit_date, last_visit_date, shop_id) VALUES ("
            f"'{cid}', {esc(phone)}, {esc(email)}, {esc(first)}, {esc(last)}, {esc(c.company)}, "
            f"{lifetime:.2f}, {len(c.appointments)}, {esc(first_visit)}, {esc(last_visit)}, 1);"
        )

        for a in c.appointments:
            if not a.date: continue
            date_seqs[a.date] += 1
            try:
                dt_obj = datetime.strptime(a.date, '%Y-%m-%d')
                bid = dt_obj.strftime('%y%m%d') + '-' + str(date_seqs[a.date]).zfill(3)
            except: continue

            cust_name = f"{first} {last}".strip() or 'Unknown'

            # Build services_json with actual field names the UI expects
            svc = {}
            if a.appointment_desc: svc['label'] = a.appointment_desc
            if a.shade: svc['filmName'] = a.shade  # e.g. "S9-15", "S5-20"
            if a.windshield: svc['windshield'] = a.windshield
            if a.tint_strip: svc['tintStrip'] = a.tint_strip
            if a.sun_roof: svc['sunRoof'] = a.sun_roof
            if a.removal: svc['removal'] = a.removal
            if a.wfi and a.wfi > 0: svc['wfi'] = a.wfi
            svc['price'] = a.total
            services_json = json.dumps([svc])

            # Clean payment field
            pay = a.payment
            if pay and len(pay) > 100: pay = pay[:100]

            book_sql.append(
                f"INSERT INTO auto_bookings (id, booking_id, customer_id, customer_name, customer_email, customer_phone, "
                f"vehicle_year, vehicle_make, vehicle_model, appointment_date, "
                f"services_json, subtotal, total_paid, balance_due, payment_method, "
                f"booking_source, status, appointment_type, service_type, notes, "
                f"calendar_event_id, shop_id, module) VALUES ("
                f"gen_random_uuid(), {esc(bid)}, '{cid}', {esc(cust_name)}, {esc(email)}, {esc(phone)}, "
                f"{a.vehicle_year or 'NULL'}, {esc(a.vehicle_make)}, {esc(a.vehicle_model)}, {esc(a.date)}, "
                f"{esc(services_json)}::jsonb, {a.total:.2f}, {a.total:.2f}, 0, {esc(pay)}, "
                f"'internal', 'completed', 'dropoff', 'tint', {esc(a.note)}, "
                f"{esc(a.event_id)}, 1, 'auto_tint');"
            )

    for dt_str, cnt in date_seqs.items():
        seq_sql.append(
            f"INSERT INTO auto_booking_sequence (sequence_date, last_number) "
            f"VALUES ('{dt_str}', {cnt}) ON CONFLICT (sequence_date) DO UPDATE SET last_number = GREATEST(auto_booking_sequence.last_number, {cnt});"
        )

    for fname, lines in [('scripts/import_customers.sql', cust_sql), ('scripts/import_bookings.sql', book_sql), ('scripts/import_sequences.sql', seq_sql)]:
        with open(fname, 'w') as f:
            if 'sequence' not in fname: f.write('BEGIN;\n')
            for l in lines: f.write(l + '\n')
            if 'sequence' not in fname: f.write('COMMIT;\n')

    print(f"\nDone:")
    print(f"  {len(cust_sql)} customers")
    print(f"  {len(book_sql)} appointments")
    print(f"  {len(seq_sql)} sequence updates")

if __name__ == '__main__':
    main()
