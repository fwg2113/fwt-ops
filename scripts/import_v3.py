#!/usr/bin/env python3
"""
FWT Import v3 -- Simple, faithful, no cleverness.
Read each row exactly as-is. Map columns directly. No parsing vehicle from deal names.
"""

import openpyxl
import json
import re
import uuid
from datetime import datetime, date
from collections import defaultdict

XLSX = 'FWT Customers & Appointments/FWT Customer _ Appointment Database.xlsx'

def norm_phone(raw):
    if raw is None: return None
    digits = re.sub(r'\D', '', str(raw).split('.')[0])
    if len(digits) == 11 and digits.startswith('1'): digits = digits[1:]
    return digits if len(digits) == 10 else None

def norm_email(raw):
    if raw is None: return None
    s = str(raw).strip().lower()
    return s if '@' in s and '.' in s else None

def s(val):
    """Convert to string or None. Skip False/True/numeric-only junk."""
    if val is None: return None
    if isinstance(val, bool): return None
    v = str(val).strip()
    if not v or v.lower() in ('none', 'false', 'true'): return None
    return v

def num(val):
    if val is None: return 0.0
    if isinstance(val, str) and not val.strip(): return 0.0
    try: return round(float(val), 2)
    except: return 0.0

def dt(val):
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, date): return val.strftime('%Y-%m-%d')
    return None

def esc(val):
    if val is None: return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # phone -> {emails: [(email, date)], first, last, company, rows: [...]}
    customers = {}

    def get(phone):
        if phone not in customers:
            customers[phone] = {'emails': [], 'first': None, 'last': None, 'company': None, 'rows': []}
        return customers[phone]

    def update_name(c, first, last):
        if first and (not c['first'] or len(first) > len(c['first'])):
            if not re.match(r'^\d', first): c['first'] = first
        if last and (not c['last'] or len(last) > len(c['last'])):
            if not re.match(r'^\d', last): c['last'] = last

    def split_name(raw):
        if not raw: return (None, None)
        n = str(raw).strip()
        if re.match(r'^\d', n): return (None, None)
        if '/' in n: n = n.split('/')[0].strip()
        parts = n.split()
        if len(parts) == 0: return (None, None)
        if len(parts) == 1: return (parts[0], None)
        return (parts[0], ' '.join(parts[1:]))

    # ==================================================================
    # 2025 Appointments
    # Cols: 0=HS, 1=Date, 2=Name, 3=Phone, 4=Email, 5=Appointment,
    #       6=Vehicle, 7=Shade, 8=WS, 9=TS, 10=SR, 11=Removal,
    #       12=Price, 13=WFI, 14=Tip, 15=TOTAL, 16=Payment, 17=Source,
    #       18=Note, 19=Company, 20=GC, 21=Invoice#, 22=RecordID, 23=EventID
    # ==================================================================
    print("2025...")
    ws = wb['2025 Appointments']
    ct = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = (list(row) + [None]*25)[:25]
        phone = norm_phone(v[3])
        if not phone: continue
        d = dt(v[1])
        if not d: continue

        first, last = split_name(v[2])
        email = norm_email(v[4])
        c = get(phone)
        update_name(c, first, last)
        if email: c['emails'].append((email, d))
        if s(v[19]): c['company'] = s(v[19])

        c['rows'].append({
            'date': d,
            'appointment': s(v[5]),   # Full appointment text as-is
            'vehicle': s(v[6]),       # Vehicle column as-is
            'shade': s(v[7]),
            'ws': s(v[8]),
            'ts': s(v[9]),
            'sr': s(v[10]),
            'removal': s(v[11]),
            'price': num(v[12]),
            'wfi': num(v[13]),
            'tip': num(v[14]),
            'total': num(v[15]),
            'payment': s(v[16]),
            'source': s(v[17]),
            'note': s(v[18]),
            'gc': s(v[20]),
            'invoice': s(v[21]),
            'event_id': s(v[23]),
        })
        ct += 1
    print(f"  {ct} rows")

    # ==================================================================
    # 2024 Appointments (same layout, skip row 2 spacer)
    # ==================================================================
    print("2024...")
    ws = wb['2024 Appointments']
    ct = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = (list(row) + [None]*24)[:24]
        phone = norm_phone(v[3])
        if not phone: continue
        d = dt(v[1])
        if not d: continue

        first, last = split_name(v[2])
        email = norm_email(v[4])
        c = get(phone)
        update_name(c, first, last)
        if email: c['emails'].append((email, d))
        if s(v[19]): c['company'] = s(v[19])

        c['rows'].append({
            'date': d,
            'appointment': s(v[5]),
            'vehicle': s(v[6]),
            'shade': s(v[7]),
            'ws': s(v[8]),
            'ts': s(v[9]),
            'sr': s(v[10]),
            'removal': None,  # 2024 doesn't have removal col (col 11 is something else)
            'price': num(v[12]),
            'wfi': num(v[13]),
            'tip': num(v[14]),
            'total': num(v[15]),
            'payment': s(v[16]),
            'source': s(v[17]),
            'note': s(v[18]),
            'gc': s(v[20]),
            'invoice': s(v[21]),
            'event_id': s(v[23]),
        })
        ct += 1
    print(f"  {ct} rows")

    # ==================================================================
    # 2020-2023 HubSpot format
    # Cols: 0=RecordId, 1=Email, 2=First, 3=Last, 4=Phone, 5=DealName,
    #       6=Film, 7=CloseDate, 8=TotalPrice, 9=GiftCert, 10=ProductType,
    #       11=Pipeline, 12=Source, 13=DupCount, 14=DealStage, 15=OrderId
    # ==================================================================
    for year in ['2020', '2021', '2022', '2023']:
        print(f"{year}...")
        ws = wb[f'{year} Appointments']
        ct = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = (list(row) + [None]*18)[:18]
            phone = norm_phone(v[4])
            if not phone: continue
            d = dt(v[7])
            if not d: continue

            first = str(v[2]).strip() if v[2] else None
            last = str(v[3]).strip() if v[3] else None
            email = norm_email(v[1])
            c = get(phone)
            update_name(c, first, last)
            if email: c['emails'].append((email, d))

            c['rows'].append({
                'date': d,
                'appointment': s(v[5]),  # Deal name as-is (e.g. "2019 Lexus LS500 S9-30 $425")
                'vehicle': None,         # No separate vehicle column
                'shade': s(v[6]),        # Film column (e.g. "S9-30")
                'ws': None, 'ts': None, 'sr': None, 'removal': None,
                'price': num(v[8]),
                'wfi': 0, 'tip': 0,
                'total': num(v[8]),      # Total = price for these sheets
                'payment': None,
                'source': s(v[12]),
                'note': None,
                'gc': s(v[9]),
                'invoice': None,
                'event_id': None,
            })
            ct += 1
        print(f"  {ct} rows")

    # ==================================================================
    # Email enrichment from Google Calendar + HubSpot list
    # ==================================================================
    print("Email enrichment...")
    ws = wb['Google Calendar Data']
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = (list(row) + [None]*14)[:14]
        phone = norm_phone(v[2])
        email = norm_email(v[3])
        if phone and email:
            c = get(phone)
            c['emails'].append((email, dt(v[4])))
            first = str(v[0]).strip() if v[0] else None
            last = str(v[1]).strip() if v[1] else None
            update_name(c, first, last)

    ws = wb['Hubspot_FWT_Email_List_with_Dup']
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = (list(row) + [None]*17)[:17]
        phone = norm_phone(v[4])
        email = norm_email(v[1])
        if phone and email:
            c = get(phone)
            c['emails'].append((email, dt(v[7])))
            first = str(v[2]).strip() if v[2] else None
            last = str(v[3]).strip() if v[3] else None
            update_name(c, first, last)

    # ==================================================================
    # Google Calendar: 2018-2019 appointments (phone + notes are the record)
    # Cols: 0=first, 1=last, 2=phone, 3=email, 4=date, 5=year, 6=services,
    #       7=film, 8=shade_front, 9=shade_rear, 10=price, 11=payment, 12=appt_type, 13=notes
    # ==================================================================
    print("Google Calendar 2018-2019...")
    ws = wb['Google Calendar Data']
    ct = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = (list(row) + [None]*14)[:14]
        phone = norm_phone(v[2])
        if not phone: continue
        d = dt(v[4])
        if not d: continue
        # Only 2018 and 2019
        if not d.startswith('2018') and not d.startswith('2019'): continue

        first = str(v[0]).strip() if v[0] else None
        last = str(v[1]).strip() if v[1] else None
        email = norm_email(v[3])
        notes_text = s(v[13])  # The calendar event title -- this IS the record
        price = num(v[10])

        c = get(phone)
        update_name(c, first, last)
        if email: c['emails'].append((email, d))

        c['rows'].append({
            'date': d,
            'appointment': notes_text,  # e.g. "2017 Ford Transit DP20/RW5 TS $267"
            'vehicle': None,
            'shade': None,
            'ws': None, 'ts': None, 'sr': None, 'removal': None,
            'price': price,
            'wfi': 0, 'tip': 0,
            'total': price,
            'payment': s(v[11]),
            'source': None,
            'note': None,
            'gc': None,
            'invoice': None,
            'event_id': None,
        })
        ct += 1
    print(f"  {ct} rows")

    wb.close()

    # Only keep customers with appointments
    customers = {p: c for p, c in customers.items() if c['rows']}

    # ==================================================================
    # Dedup: phone + date + total (exact match)
    # ==================================================================
    print("Dedup...")
    total_before = sum(len(c['rows']) for c in customers.values())
    for p, c in customers.items():
        seen = {}
        unique = []
        for r in c['rows']:
            key = f"{r['date']}|{r['total']}"
            if key not in seen:
                seen[key] = r
                unique.append(r)
            else:
                # Merge: prefer the one with more fields filled
                existing = seen[key]
                for k in ['appointment', 'vehicle', 'shade', 'ws', 'ts', 'sr', 'payment', 'note', 'event_id']:
                    if r.get(k) and not existing.get(k):
                        existing[k] = r[k]
        c['rows'] = unique
    total_after = sum(len(c['rows']) for c in customers.values())
    print(f"  {total_before} -> {total_after} ({total_before - total_after} dupes)")

    # ==================================================================
    # Generate SQL
    # ==================================================================
    print(f"\nGenerating SQL: {len(customers)} customers, {total_after} appointments...")

    cust_lines = []
    book_lines = []
    seq_counts = defaultdict(int)

    for phone, c in customers.items():
        cid = str(uuid.uuid4())

        # Best email = most recent
        best_email = None
        if c['emails']:
            best_email = sorted(c['emails'], key=lambda x: x[1] or '0000', reverse=True)[0][0]

        first = c['first'] or ''
        last = c['last'] or ''
        name = f"{first} {last}".strip() or 'Unknown'

        dates = [r['date'] for r in c['rows']]
        totals = [r['total'] for r in c['rows'] if r['total'] > 0]

        cust_lines.append(
            f"INSERT INTO customers (id, phone, email, first_name, last_name, company_name, "
            f"lifetime_spend, visit_count, first_visit_date, last_visit_date, shop_id) VALUES ("
            f"'{cid}', {esc(phone)}, {esc(best_email)}, {esc(first)}, {esc(last)}, {esc(c['company'])}, "
            f"{sum(totals):.2f}, {len(c['rows'])}, {esc(min(dates))}, {esc(max(dates))}, 1);"
        )

        for r in c['rows']:
            seq_counts[r['date']] += 1
            seq = seq_counts[r['date']]
            try:
                bid = datetime.strptime(r['date'], '%Y-%m-%d').strftime('%y%m%d') + '-' + str(seq).zfill(3)
            except:
                continue

            # Parse vehicle string into year/make/model
            vy, vmake, vmodel = 'NULL', 'NULL', 'NULL'
            veh = r.get('vehicle')
            if veh:
                m = re.match(r'(\d{4})\s+(.+)', veh)
                if m:
                    vy = m.group(1)
                    parts = m.group(2).split()
                    vmake = esc(parts[0]) if parts else 'NULL'
                    vmodel = esc(' '.join(parts[1:])) if len(parts) > 1 else 'NULL'
                else:
                    parts = veh.split()
                    vmake = esc(parts[0]) if parts else 'NULL'
                    vmodel = esc(' '.join(parts[1:])) if len(parts) > 1 else 'NULL'

            # Build services_json -- just store everything we have
            svc = {}
            if r['appointment']: svc['label'] = r['appointment']
            if r['shade']: svc['filmName'] = r['shade']
            if r['ws']: svc['windshield'] = r['ws']
            if r['ts']: svc['tintStrip'] = r['ts']
            if r['sr']: svc['sunRoof'] = r['sr']
            if r['removal']: svc['removal'] = r['removal']
            if r['wfi'] and r['wfi'] > 0: svc['wfi'] = r['wfi']
            if r['tip'] and r['tip'] > 0: svc['tip'] = r['tip']
            svc['price'] = r['total']
            svc_json = json.dumps([svc])

            pay = r.get('payment')
            if pay and len(pay) > 100: pay = pay[:100]

            book_lines.append(
                f"INSERT INTO auto_bookings (id, booking_id, customer_id, customer_name, customer_email, customer_phone, "
                f"vehicle_year, vehicle_make, vehicle_model, appointment_date, "
                f"services_json, subtotal, total_paid, balance_due, payment_method, "
                f"booking_source, status, appointment_type, service_type, notes, "
                f"calendar_event_id, shop_id, module) VALUES ("
                f"gen_random_uuid(), {esc(bid)}, '{cid}', {esc(name)}, {esc(best_email)}, {esc(phone)}, "
                f"{vy}, {vmake}, {vmodel}, {esc(r['date'])}, "
                f"{esc(svc_json)}::jsonb, {r['total']:.2f}, {r['total']:.2f}, 0, {esc(pay)}, "
                f"'internal', 'completed', 'dropoff', 'tint', {esc(r.get('note'))}, "
                f"{esc(r.get('event_id'))}, 1, 'auto_tint');"
            )

    seq_lines = []
    for d, cnt in seq_counts.items():
        seq_lines.append(f"INSERT INTO auto_booking_sequence (sequence_date, last_number) VALUES ('{d}', {cnt}) ON CONFLICT (sequence_date) DO UPDATE SET last_number = GREATEST(auto_booking_sequence.last_number, {cnt});")

    for fname, lines in [('scripts/import_customers.sql', cust_lines), ('scripts/import_bookings.sql', book_lines), ('scripts/import_sequences.sql', seq_lines)]:
        with open(fname, 'w') as f:
            if 'sequence' not in fname: f.write('BEGIN;\n')
            for l in lines: f.write(l + '\n')
            if 'sequence' not in fname: f.write('COMMIT;\n')

    print(f"Done: {len(cust_lines)} customers, {len(book_lines)} bookings")

if __name__ == '__main__':
    main()
