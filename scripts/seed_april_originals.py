#!/usr/bin/env python3
"""
Seed April 2026 originals (Session 10, hybrid plan)
====================================================
Creates auto_bookings rows for all 52 April 2026 invoices using ONLY the
ORIGINAL booking data from the BookingLog spreadsheet. No line item inference,
no discount back-solving, no menu price lookups, no upsell math.

Each row gets:
  - Customer info, vehicle, date, time, duration (from BookingLog)
  - Original services_json with the ONE original service the customer booked
  - Original $ subtotal (from BookingLog)
  - starting_total_override = same value (locks the original for upsell math)
  - Deposit (from BookingLog, may be $0)
  - Status = 'completed' (work done, ready for the team to invoice)
  - import_source = 'historic_import' (so the wipe step can find them)

After this script runs, Joe walks each appointment through the live UI:
  1. Click Edit → add the actual services performed (per spreadsheet)
  2. Click Invoice → counter checkout opens
  3. Pick warranty + discounts + payment method
  4. Pay → live system writes correct upsell math via CollectPaymentModal

NO API routes called. NO SMS/email/calendar/cron triggered. Direct REST writes.

Usage:
  python3 scripts/seed_april_originals.py --show 260401-001
      Print the proposed booking for one row, no DB writes

  python3 scripts/seed_april_originals.py --wipe
      Delete all April historic_import bookings + linked docs/lines/payments

  python3 scripts/seed_april_originals.py --seed-one 260401-001
      Insert one row (after wipe)

  python3 scripts/seed_april_originals.py --seed-all
      Insert all 52 April rows (after wipe)
"""

import openpyxl
import json
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime, date
from typing import Optional

# ============================================================================
# CONFIG
# ============================================================================
SUPABASE_URL = "https://jyrgqbhpdbaiedfgoyse.supabase.co"
SERVICE_KEY: Optional[str] = None
SHOP_ID = 1

FINANCIAL_XLSX = "legacy/spreadsheets/2026 Financial Core – FWT & FWG (1).xlsx"
BOOKINGLOG_XLSX = "legacy/spreadsheets/FWT_Booking_Spec_FINAL (1).xlsx"

# Service labels — pulled from the live auto_services table.
# Used only to write a human-readable label on the booking's services_json
# entry; the live EditAppointmentModal will read these the same way it reads
# any new appointment.
LABEL_MAP = {
    'FULL_SIDES': 'Full (Sides & Rear)',
    'TWO_FRONT_DOORS': '2 Front Doors Only',
    'FULL_WS': 'Full Windshield',
    'SUN_STRIP': 'Sun Strip',
    'SUNROOF_SINGLE': 'Single Sunroof',
    'SUNROOF_PANO': 'Full Panoramic Sunroof',
    'REMOVAL_FULL': 'Removal - Full (Sides & Rear)',
}

# Film abbrev → film_id (verified live in session 10)
FILM_BY_NAME = {
    'Black': 5,
    'Black Ceramic': 2,
    'Ceramic i3': 3,
    'Ceramic i3+': 4,
}
FILM_ABBREV = {
    5: 'BLK',
    2: 'BC',
    3: 'i3',
    4: 'i3+',
}


# ============================================================================
# ENV
# ============================================================================
def load_env():
    global SERVICE_KEY
    with open(".env.local") as f:
        for line in f:
            if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
                SERVICE_KEY = line.split("=", 1)[1].strip()
                return
    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY not found in .env.local")


# ============================================================================
# SUPABASE REST CLIENT
# ============================================================================
def sb(method, path, body=None, params=None):
    from urllib.parse import quote
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        qs = "&".join(f"{k}={quote(str(v), safe='().,*:')}" for k, v in params.items())
        url = f"{url}?{qs}"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body, default=str).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text else None
    except urllib.error.HTTPError as e:
        body_text = e.read().decode()
        print(f"HTTP {e.code}: {body_text}", file=sys.stderr)
        if body is not None:
            print(f"  body: {json.dumps(body, default=str)[:600]}", file=sys.stderr)
        raise


# ============================================================================
# HELPERS
# ============================================================================
def normalize_phone(raw):
    if raw is None:
        return None
    s = str(raw).split('.')[0].strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    return digits if len(digits) == 10 else None


def normalize_email(raw):
    if raw is None:
        return None
    s = str(raw).strip().lower()
    return s if '@' in s and '.' in s else None


def safe_str(v):
    return None if v is None else str(v).strip()


def to_num(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def to_time_str(v):
    if v is None:
        return None
    if hasattr(v, 'hour'):
        return v.strftime('%H:%M:%S')
    return str(v)


def to_local_noon_iso(v):
    """
    Returns an ISO timestamp at 17:00 UTC (1pm EDT / noon EST) for the given
    date. Without this, a bare date like '2026-04-01' is parsed as midnight UTC,
    which is 8pm March 31 EDT — and the dashboard's invoicing page filters it
    as March instead of April.
    """
    d = None
    if isinstance(v, datetime):
        d = v.date()
    elif isinstance(v, date):
        d = v
    else:
        s = str(v).strip()
        if not s:
            return None
        try:
            d = datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            return s
    return f"{d.isoformat()}T17:00:00+00:00"


def round2(x):
    return round(float(x) + 1e-9, 2)


# ============================================================================
# LOAD APRIL ROWS FROM 2026FWT (used to know WHICH event_ids to seed,
# and to provide the InvoiceNum that becomes the booking_id)
# ============================================================================
def load_april_invoice_nums():
    wb = openpyxl.load_workbook(FINANCIAL_XLSX, data_only=True)
    ws = wb['2026FWT']
    out = []
    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        d = row[0]
        if not d or not hasattr(d, 'year') or d.year != 2026 or d.month != 4:
            continue
        invoice_num = safe_str(row[29])
        event_id = safe_str(row[30])
        if not invoice_num:
            continue
        out.append({
            'invoice_num': invoice_num,
            'event_id': event_id,
            'sheet_date': d,
        })
    return out


# ============================================================================
# LOAD BOOKINGLOG indexed by event_id
# ============================================================================
_booking_log_by_event = None


def load_booking_log():
    global _booking_log_by_event
    if _booking_log_by_event is not None:
        return _booking_log_by_event
    wb = openpyxl.load_workbook(BOOKINGLOG_XLSX, data_only=True)
    ws = wb['BookingLog']
    out = {}
    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        event = safe_str(row[25])
        if not event:
            continue
        try:
            services = json.loads(row[16]) if row[16] else []
        except (TypeError, ValueError):
            services = []
        out[event] = {
            'booking_id': safe_str(row[0]),
            'created_at': row[1],
            'booking_source': safe_str(row[2]),
            'service_type': safe_str(row[3]),
            'appointment_type': safe_str(row[4]),
            'emoji_marker': safe_str(row[5]),
            'name': safe_str(row[6]),
            'email': safe_str(row[7]),
            'phone': safe_str(row[8]),
            'vehicle_year': int(row[9]) if row[9] else None,
            'vehicle_make': safe_str(row[10]),
            'vehicle_model': safe_str(row[11]),
            'class_keys': safe_str(row[12]),
            'date': row[13],
            'time': row[14],
            'duration': int(row[15]) if row[15] else None,
            'services_json_raw': services,
            'subtotal': to_num(row[18]),
            'discount_code': safe_str(row[19]),
            'discount_amount': to_num(row[22]),
            'deposit_paid': to_num(row[23]),
            'balance_due': to_num(row[24]),
            'status': safe_str(row[27]),
            'notes': safe_str(row[30]),
        }
    _booking_log_by_event = out
    return out


# ============================================================================
# BUILD ONE BOOKING RECORD FROM A BookingLog ENTRY + spreadsheet invoice_num
# ============================================================================
def build_booking_record(invoice_num, bl_match):
    """
    Builds the auto_bookings INSERT body for one row.
    NO inference, NO back-solving — every value comes either from BookingLog
    directly or is a fixed constant (status, module, import_source).
    """
    # Map raw booking_source to allowed enum
    raw_src = (bl_match.get('booking_source') or '').lower()
    if raw_src in ('nodep', 'shopify', 'online'):
        booking_source = 'online'
    elif raw_src == 'phone':
        booking_source = 'phone'
    elif raw_src in ('gc', 'gift_certificate'):
        booking_source = 'gc'
    elif raw_src == 'sameday':
        booking_source = 'sameday'
    else:
        booking_source = 'internal'

    # Map raw appointment_type to allowed enum
    raw_appt = (bl_match.get('appointment_type') or '').lower()
    if raw_appt in ('dropoff', 'waiting', 'headsup_30', 'headsup_60'):
        appt_type = raw_appt
    else:
        appt_type = 'dropoff'

    # Convert the BookingLog services_json (snake_case) to live TintServiceLine
    # shape (camelCase) so the dashboard renders it correctly. NOTHING is
    # inferred — the same data, just renamed fields and a label added.
    services_live = []
    for svc in bl_match.get('services_json_raw', []):
        film_name = svc.get('film_type')
        film_id = FILM_BY_NAME.get(film_name) if film_name else None
        film_abbrev = FILM_ABBREV.get(film_id) if film_id else None
        sk = svc.get('service_key', '')
        services_live.append({
            'serviceKey': sk,
            'label': LABEL_MAP.get(sk, sk),
            'filmId': film_id,
            'filmName': film_name,
            'filmAbbrev': film_abbrev,
            'shadeFront': svc.get('shade_front') or None,
            'shadeRear': svc.get('shade_rear') or None,
            'shade': svc.get('shade') or svc.get('shade_front') or svc.get('shade_rear'),
            'price': float(svc.get('base_price') or 0),
            'discountAmount': 0,
            'duration': 0,
        })

    starting_total = round2(bl_match.get('subtotal', 0))
    deposit = round2(bl_match.get('deposit_paid', 0))
    discount = round2(bl_match.get('discount_amount', 0))
    appt_date = bl_match['date'].date() if hasattr(bl_match['date'], 'date') else bl_match['date']

    return {
        'shop_id': SHOP_ID,
        'booking_id': invoice_num,
        'booking_source': booking_source,
        'appointment_type': appt_type,
        'service_type': 'tint',
        'emoji_marker': bl_match.get('emoji_marker'),
        'customer_name': bl_match.get('name'),
        'customer_phone': normalize_phone(bl_match.get('phone')),
        'customer_email': normalize_email(bl_match.get('email')),
        'vehicle_year': bl_match.get('vehicle_year'),
        'vehicle_make': bl_match.get('vehicle_make'),
        'vehicle_model': bl_match.get('vehicle_model'),
        'class_keys': bl_match.get('class_keys'),  # raw pipe-separated string from BookingLog
        'appointment_date': to_date_str(appt_date),
        'appointment_time': to_time_str(bl_match.get('time')),
        'duration_minutes': bl_match.get('duration'),
        'services_json': services_live,
        'subtotal': starting_total,
        'starting_total_override': starting_total,  # locks the original for upsell math
        'discount_code': bl_match.get('discount_code'),
        'discount_amount': discount,
        'deposit_paid': deposit,
        # balance_due must subtract BOTH the deposit AND any pre-applied
        # discount (e.g. gift certificate redemption from BookingLog).
        'balance_due': round2(starting_total - deposit - discount),
        'total_paid': 0,
        'payment_method': None,
        'calendar_event_id': None,  # Joe will let the live system create cal events forward
        'status': 'completed',  # work done, ready to be invoiced via the live UI
        'module': 'auto_tint',
        'import_source': 'historic_import',
        'notes': None,
        # Pre-set so the review_request cron skips this row even if Joe forgets
        # to keep the toggle off (defense in depth — toggle is the primary guard).
        'review_request_sent_at': datetime.utcnow().isoformat() + '+00:00',
    }


# ============================================================================
# WIPE
# ============================================================================
def wipe_april():
    """
    Delete all April 2026 historic_import bookings + linked documents.
    Same logic as scripts/import_april_2026.py wipe — finds rows by both
    document.created_at and auto_bookings.booking_id matching April invoices.
    """
    print("=== WIPING APRIL 2026 historic_import data ===")

    # Pass 1: documents in April
    docs = sb("GET", "documents", params={
        "select": "id,doc_number",
        "shop_id": f"eq.{SHOP_ID}",
        "import_source": "eq.historic_import",
        "created_at": "gte.2026-04-01",
        "and": "(created_at.lt.2026-05-01)",
    }) or []
    print(f"Found {len(docs)} April documents to wipe")

    if docs:
        doc_ids = [d['id'] for d in docs]
        print("  Deleting document_payments...")
        sb("DELETE", "document_payments", params={
            "document_id": f"in.({','.join(doc_ids)})"
        })
        print("  Deleting document_line_items...")
        sb("DELETE", "document_line_items", params={
            "document_id": f"in.({','.join(doc_ids)})"
        })
        print("  Unlinking auto_bookings.document_id...")
        sb("PATCH", "auto_bookings",
           body={"document_id": None},
           params={"document_id": f"in.({','.join(doc_ids)})"})
        print("  Deleting documents...")
        sb("DELETE", "documents", params={
            "id": f"in.({','.join(doc_ids)})"
        })

    # Pass 2: orphaned bookings by booking_id
    april_invs = [r['invoice_num'] for r in load_april_invoice_nums()]
    if april_invs:
        in_list = ','.join(f'"{v}"' for v in april_invs)
        bookings = sb("GET", "auto_bookings", params={
            "select": "id,booking_id,customer_name,document_id",
            "shop_id": f"eq.{SHOP_ID}",
            "import_source": "eq.historic_import",
            "booking_id": f"in.({in_list})",
        }) or []
        if bookings:
            print(f"Found {len(bookings)} April-numbered auto_bookings to wipe")
            booking_uuids = [b['id'] for b in bookings]
            sb("DELETE", "auto_bookings", params={
                "id": f"in.({','.join(booking_uuids)})"
            })

    print("=== WIPE COMPLETE ===")


# ============================================================================
# DISPLAY
# ============================================================================
def show_record(invoice_num, bl_match, record):
    print()
    print("=" * 78)
    print(f"PROPOSED SEED BOOKING FOR {invoice_num}")
    print("=" * 78)
    print()
    print(f"BookingLog source: {bl_match['booking_id']} | "
          f"{bl_match['vehicle_year']} {bl_match['vehicle_make']} {bl_match['vehicle_model']} "
          f"| original ${bl_match['subtotal']}")
    print()
    print("auto_bookings INSERT body:")
    for k, v in record.items():
        if k == 'services_json' and v:
            print(f"  {k}:")
            for svc in v:
                print(f"    - {json.dumps(svc, default=str)}")
        else:
            print(f"  {k}: {v!r}")
    print()


# ============================================================================
# INSERT
# ============================================================================
def insert_record(record):
    print(f"  Inserting auto_bookings {record['booking_id']}...")
    resp = sb("POST", "auto_bookings", body=record)
    booking_uuid = resp[0]['id']
    print(f"    → {booking_uuid}")
    return booking_uuid


# ============================================================================
# MAIN
# ============================================================================
def main():
    load_env()
    args = sys.argv[1:]

    if "--show" in args:
        invoice = args[args.index("--show") + 1]
        load_booking_log()
        rows = load_april_invoice_nums()
        match = next((r for r in rows if r['invoice_num'] == invoice), None)
        if not match:
            print(f"No April row with invoice {invoice}")
            return
        bl_match = _booking_log_by_event.get(match['event_id'])
        if not bl_match:
            print(f"No BookingLog match for event_id {match['event_id']}")
            return
        record = build_booking_record(invoice, bl_match)
        show_record(invoice, bl_match, record)
        return

    if "--wipe" in args:
        if "--yes" not in args:
            resp = input("Wipe ALL April 2026 historic_import bookings + docs? [yes/no] ")
            if resp.strip().lower() != 'yes':
                print("Aborted.")
                return
        wipe_april()
        return

    if "--seed-one" in args:
        invoice = args[args.index("--seed-one") + 1]
        load_booking_log()
        rows = load_april_invoice_nums()
        match = next((r for r in rows if r['invoice_num'] == invoice), None)
        if not match:
            print(f"No April row with invoice {invoice}")
            return
        bl_match = _booking_log_by_event.get(match['event_id'])
        if not bl_match:
            print(f"No BookingLog match for event_id {match['event_id']}")
            return
        record = build_booking_record(invoice, bl_match)
        show_record(invoice, bl_match, record)
        if "--yes" not in args:
            resp = input("Insert this row? [yes/no] ")
            if resp.strip().lower() != 'yes':
                print("Aborted.")
                return
        insert_record(record)
        print("DONE.")
        return

    if "--seed-all" in args:
        load_booking_log()
        rows = load_april_invoice_nums()
        # Pre-validate every row has a BookingLog match
        unmatched = [r['invoice_num'] for r in rows if not _booking_log_by_event.get(r['event_id'] or '')]
        if unmatched:
            print(f"WARNING: {len(unmatched)} rows have no BookingLog match: {unmatched}")
            print("Aborting. Fix matching before running --seed-all.")
            return

        # Query DB for already-seeded booking_ids so we can skip them.
        # Lets you safely re-run --seed-all after manually seeding a few rows
        # for testing — only the missing ones are inserted.
        all_invoice_nums = [r['invoice_num'] for r in rows]
        in_list = ','.join(f'"{v}"' for v in all_invoice_nums)
        existing = sb("GET", "auto_bookings", params={
            "select": "booking_id",
            "shop_id": f"eq.{SHOP_ID}",
            "import_source": "eq.historic_import",
            "booking_id": f"in.({in_list})",
        }) or []
        existing_ids = {b['booking_id'] for b in existing}
        to_seed = [r for r in rows if r['invoice_num'] not in existing_ids]

        print(f"Loaded {len(rows)} April rows, all matched to BookingLog.")
        print(f"Already seeded: {len(existing_ids)} ({sorted(existing_ids)})")
        print(f"To insert: {len(to_seed)}")
        if not to_seed:
            print("Nothing to do. All April rows are already seeded.")
            return
        if "--yes" not in args:
            resp = input(f"Insert {len(to_seed)} rows? [yes/no] ")
            if resp.strip().lower() != 'yes':
                print("Aborted.")
                return
        inserted = 0
        for r in to_seed:
            bl_match = _booking_log_by_event.get(r['event_id'])
            record = build_booking_record(r['invoice_num'], bl_match)
            print(f"\n--- {r['invoice_num']} {bl_match['name']} ---")
            insert_record(record)
            inserted += 1
        print(f"\nALL DONE. Inserted {inserted} bookings (skipped {len(existing_ids)}).")
        return

    print(__doc__)


if __name__ == '__main__':
    main()
