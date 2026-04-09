#!/usr/bin/env python3
"""
April 2026 Re-Import (Session 10)
==================================
Re-imports April 2026 (52 rows) from 2026FWT spreadsheet, one row at a time
or in batch, with REAL per-line YMM pricing instead of fake even-split prices.

Approach:
  - Line items at full menu price (looked up in auto_pricing)
  - When line items sum > spreadsheet sub, the variance is recorded as a custom
    discount entry in applied_discounts (e.g. friend/family favor)
  - The transaction ledger is NOT touched (already correct)
  - NO API routes called, NO SMS/email/calendar/notifications
  - Direct Supabase REST writes only

Usage:
  python3 scripts/import_april_2026.py --show 260401-001
      Print the proposed inserts for one row, no DB writes

  python3 scripts/import_april_2026.py --wipe
      Delete all 52 April historic_import documents + bookings (after confirm)

  python3 scripts/import_april_2026.py --import 260401-001
      Insert one row (assumes wipe already happened for that row)

  python3 scripts/import_april_2026.py --import-all
      Insert all 52 April rows (after wipe)
"""

import openpyxl
import json
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime, date
from typing import Optional

# ============================================================================
# CONFIG
# ============================================================================
SUPABASE_URL = "https://jyrgqbhpdbaiedfgoyse.supabase.co"
SERVICE_KEY: Optional[str] = None
SHOP_ID = 1

FINANCIAL_XLSX = "legacy/spreadsheets/2026 Financial Core – FWT & FWG (1).xlsx"
BOOKINGLOG_XLSX = "legacy/spreadsheets/FWT_Booking_Spec_FINAL (1).xlsx"

# Static maps from session 9 audit (verified live in session 10)
FILM_BY_ABBREV = {
    "BLK": {"id": 5, "name": "Black", "abbrev": "BLK"},
    "BC":  {"id": 2, "name": "Black Ceramic", "abbrev": "BC"},
    "i3":  {"id": 3, "name": "Ceramic i3", "abbrev": "i3"},
    "i3+": {"id": 4, "name": "Ceramic i3+", "abbrev": "i3+"},
}

# warranty_product_options ids (verified live in session 10)
NFW_BY_ABBREV = {
    "NF2F":   {"option_id": 2, "warranty_product_id": 1, "name": "No Fault Warranty",
               "option_name": "No Fault - 2 Front Doors", "price": 25},
    "NFWS":   {"option_id": 4, "warranty_product_id": 1, "name": "No Fault Warranty",
               "option_name": "No Fault - Windshield", "price": 50},
    "NFF":    {"option_id": 3, "warranty_product_id": 1, "name": "No Fault Warranty",
               "option_name": "No Fault - Full", "price": 75},
    "NF2FWS": {"option_id": 5, "warranty_product_id": 1, "name": "No Fault Warranty",
               "option_name": "No Fault - 2FD + Windshield", "price": 75},
    "NFFWS":  {"option_id": 6, "warranty_product_id": 1, "name": "No Fault Warranty",
               "option_name": "No Fault - Full + Windshield", "price": 125},
}

# checkout_discount_types ids (verified live in session 10)
CASH_DISCOUNT_TYPE = {
    "discount_type_id": 5,
    "name": "Cash Discount",
    "discount_type": "percent",
    "discount_value": 5,
    "applies_to": "balance_due",
    "include_warranty": False,
}
MILITARY_25_DISCOUNT_TYPE = {
    "discount_type_id": 1,
    "name": "Military Discount",
    "discount_type": "dollar",
    "discount_value": 25,
    "applies_to": "balance_due",
    "include_warranty": False,
}
MILITARY_10_DISCOUNT_TYPE = {
    "discount_type_id": 2,
    "name": "Military Discount",
    "discount_type": "dollar",
    "discount_value": 10,
    "applies_to": "balance_due",
    "include_warranty": False,
}

# ============================================================================
# ENV
# ============================================================================
def load_env():
    global SERVICE_KEY
    with open(".env.local") as f:
        for line in f:
            if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
                SERVICE_KEY = line.split("=", 1)[1].strip()
                return
    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY not found in .env.local")


# ============================================================================
# SUPABASE REST CLIENT
# ============================================================================
def sb(method, path, body=None, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        # URL-encode each value (so spaces, parentheses, etc. don't break the URL).
        # quote_plus is wrong here because PostgREST treats '+' as ' ' (form-encoded);
        # use quote with safe='' to encode spaces as %20 and leave PostgREST operator
        # punctuation (like '.', '(', ')') alone in operator-style values.
        from urllib.parse import quote
        qs = "&".join(f"{k}={quote(str(v), safe='().,*:')}" for k, v in params.items())
        url = f"{url}?{qs}"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body, default=str).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text else None
    except urllib.error.HTTPError as e:
        body_text = e.read().decode()
        print(f"HTTP {e.code}: {body_text}", file=sys.stderr)
        if body is not None:
            print(f"  body: {json.dumps(body, default=str)[:600]}", file=sys.stderr)
        raise


# ============================================================================
# HELPERS
# ============================================================================
def normalize_phone(raw):
    if raw is None:
        return None
    s = str(raw).split('.')[0].strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    return digits if len(digits) == 10 else None


def normalize_email(raw):
    if raw is None:
        return None
    s = str(raw).strip().lower()
    return s if '@' in s and '.' in s else None


def safe_str(v):
    return None if v is None else str(v).strip()


def to_num(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def to_local_noon_iso(v):
    """
    Convert a date-like value to an ISO timestamp at 17:00 UTC (1pm EDT / noon EST).
    This places the timestamp safely in the middle of the day for any continental
    US timezone, so month-based filters using local time always classify it under
    the correct calendar month. Without this, a bare date like '2026-04-01' is
    parsed as midnight UTC, which is 8pm March 31 EDT — and the dashboard's
    invoicing page filters it as March instead of April.
    """
    d = None
    if isinstance(v, datetime):
        d = v.date()
    elif isinstance(v, date):
        d = v
    else:
        s = str(v).strip()
        if not s:
            return None
        try:
            d = datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            return s
    return f"{d.isoformat()}T17:00:00+00:00"


def to_time_str(v):
    if v is None:
        return None
    if hasattr(v, 'hour'):
        return v.strftime('%H:%M:%S')
    return str(v)


def round2(x):
    return round(float(x) + 1e-9, 2)


# ============================================================================
# FILM + SHADE PARSING
# ============================================================================
# Service column patterns:
#   FULL: "BC 30%" or "BC 30%/40%" (front/rear) or "i3 15%/56%"
#   2FD:  "BC 30%" or "BC MATCH%"  (MATCH = match the rear shade)
#   WS:   "i3 70%"  (single shade)
#   TS:   always "BLK 5%"
#   SR:   "SRx2 BC 30%" (panoramic) or "SR BC 30%" (single) or just "BC 30%"
#   Removal: dollar amount as text or number
def parse_film_shade(raw, allow_match=False):
    """
    Parse a service column value.
    Returns (film_abbrev, shade_single, shade_front, shade_rear).
    For single-shade services, shade_single is set and front/rear are None.
    For dual-shade ("30%/40%"), front/rear are set and shade_single is None.
    """
    if raw is None:
        return None, None, None, None
    s = str(raw).strip()
    if not s:
        return None, None, None, None
    # Strip SR / SRx2 prefix if present (handled separately)
    s = re.sub(r'^SR(?:x2)?\s+', '', s, flags=re.IGNORECASE).strip()
    # Try "BC 30%" or "i3+ 30%/40%" or "BC MATCH%"
    m = re.match(r'^(BC|BLK|i3\+|i3)\s+(.+)$', s)
    if not m:
        return None, s if s else None, None, None
    abbrev = m.group(1)
    shade_text = m.group(2).strip()
    # Dual shade "30%/40%" → shade_front, shade_rear
    if '/' in shade_text:
        parts = [p.strip() for p in shade_text.split('/', 1)]
        if len(parts) == 2:
            return abbrev, None, parts[0], parts[1]
    return abbrev, shade_text, None, None


def is_panoramic(raw):
    if raw is None:
        return False
    return bool(re.match(r'^SRx2\b', str(raw).strip(), re.IGNORECASE))


# ============================================================================
# DB LOOKUPS
# ============================================================================
_vehicle_cache = {}
_pricing_cache = {}
# auto_vehicle_classes lookup, populated lazily on first call
_class_meta_cache = None


def load_class_meta():
    """
    Load auto_vehicle_classes once. Returns dict { class_key: { is_add_fee, add_fee_amount } }.
    Used to detect ADD_FEE_* surcharges and skip non-priced classes.
    """
    global _class_meta_cache
    if _class_meta_cache is not None:
        return _class_meta_cache
    rows = sb("GET", "auto_vehicle_classes",
              params={"select": "class_key,name,is_add_fee,add_fee_amount"}) or []
    _class_meta_cache = {
        r['class_key']: {
            'name': r.get('name'),
            'is_add_fee': r.get('is_add_fee', False),
            'add_fee_amount': float(r.get('add_fee_amount') or 0),
        }
        for r in rows
    }
    return _class_meta_cache


def lookup_vehicle(year, make, model):
    """
    Returns (vehicle_id, class_keys_list) or (None, []).
    The list preserves insertion order from auto_vehicles.class_keys —
    matches the live resolveClassKey behavior (first match wins).
    """
    if not (year and make and model):
        return None, []
    key = (int(year), str(make).strip().lower(), str(model).strip().lower())
    if key in _vehicle_cache:
        return _vehicle_cache[key]
    params = {
        "select": "id,class_keys,window_count",
        "make": f"ilike.{make}",
        "model": f"ilike.{model}",
        "year_start": f"lte.{year}",
        "year_end": f"gte.{year}",
    }
    rows = sb("GET", "auto_vehicles", params=params)
    if not rows:
        _vehicle_cache[key] = (None, [])
        return None, []
    row = rows[0]
    class_keys = row.get("class_keys") or []
    _vehicle_cache[key] = (row["id"], class_keys)
    return row["id"], class_keys


def lookup_price(class_keys_list, service_key, film_id):
    """
    Look up auto_pricing for the first class in the list that has a matching row.
    Falls back to class_key='*' for universal services (sunroofs, windshields, etc.).
    Matches the live src/app/components/booking/pricing.ts:resolveClassKey behavior.
    Returns (price, duration, matched_class_key) or (None, None, None).
    """
    if isinstance(class_keys_list, str):
        class_keys_list = [class_keys_list]
    if not class_keys_list:
        class_keys_list = []

    # Try each class in order, then wildcard '*'
    candidates = list(class_keys_list) + ["*"]
    for ck in candidates:
        if ck is None:
            continue
        cache_key = (ck, service_key, film_id)
        if cache_key in _pricing_cache:
            cached = _pricing_cache[cache_key]
            if cached[0] is not None:
                return cached[0], cached[1], ck
            continue

        params = {
            "select": "price,duration_minutes",
            "shop_id": f"eq.{SHOP_ID}",
            "class_key": f"eq.{ck}",
            "service_key": f"eq.{service_key}",
        }
        if film_id is not None:
            params["film_id"] = f"eq.{film_id}"
        else:
            params["film_id"] = "is.null"
        rows = sb("GET", "auto_pricing", params=params)
        if rows:
            price = float(rows[0]["price"])
            duration = rows[0].get("duration_minutes")
            _pricing_cache[cache_key] = (price, duration)
            return price, duration, ck
        _pricing_cache[cache_key] = (None, None)
    return None, None, None


def get_add_fee(class_keys_list):
    """Sum add_fee_amount for any ADD_FEE_* classes on the vehicle."""
    if not class_keys_list:
        return 0.0
    meta = load_class_meta()
    total = 0.0
    for ck in class_keys_list:
        m = meta.get(ck)
        if m and m['is_add_fee']:
            total += m['add_fee_amount']
    return total


# ============================================================================
# BOOKINGLOG LOADER (for original booking lookup)
# ============================================================================
_booking_log_by_event = None


def load_booking_log():
    global _booking_log_by_event
    if _booking_log_by_event is not None:
        return _booking_log_by_event
    wb = openpyxl.load_workbook(BOOKINGLOG_XLSX, data_only=True)
    ws = wb['BookingLog']
    out = {}
    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        event = safe_str(row[25])
        if not event:
            continue
        try:
            services = json.loads(row[16]) if row[16] else []
        except (TypeError, ValueError):
            services = []
        for svc in services:
            if not svc.get('shade'):
                svc['shade'] = svc.get('shade_front') or svc.get('shade_rear')
        out[event] = {
            'booking_id': safe_str(row[0]),
            'created_at': row[1],
            'booking_source': safe_str(row[2]),
            'service_type': safe_str(row[3]),
            'appointment_type': safe_str(row[4]),
            'emoji_marker': safe_str(row[5]),
            'name': safe_str(row[6]),
            'email': safe_str(row[7]),
            'phone': safe_str(row[8]),
            'vehicle_year': int(row[9]) if row[9] else None,
            'vehicle_make': safe_str(row[10]),
            'vehicle_model': safe_str(row[11]),
            'class_keys': safe_str(row[12]),
            'date': row[13],
            'time': row[14],
            'duration': int(row[15]) if row[15] else None,
            'services_json': services,
            'subtotal': to_num(row[18]),
            'discount_code': safe_str(row[19]),
            'discount_amount': to_num(row[22]),
            'deposit_paid': to_num(row[23]),
            'balance_due': to_num(row[24]),
            'status': safe_str(row[27]),
            'notes': safe_str(row[30]),
        }
    _booking_log_by_event = out
    return out


# ============================================================================
# ROW PARSER
# ============================================================================
# Service labels — pulled from auto_services audit (session 10).
# These match the live booking flow exactly so the dashboard renders these
# the same way it renders any new appointment.
LABEL_MAP = {
    'FULL_SIDES': 'Full (Sides & Rear)',
    'TWO_FRONT_DOORS': '2 Front Doors Only',
    'FULL_WS': 'Full Windshield',
    'SUN_STRIP': 'Sun Strip',
    'SUNROOF_SINGLE': 'Single Sunroof',
    'SUNROOF_PANO': 'Full Panoramic Sunroof',
    'REMOVAL_FULL': 'Removal - Full (Sides & Rear)',
    'REMOVAL_2FD': 'Removal - 2 Front Doors',
    'REMOVAL_WS': 'Removal - Windshield',
    'REMOVAL_SUNSTRIP': 'Removal - Sun Strip',
    'REMOVAL_REAR': 'Removal - Rear Window',
    'ALACARTE_DFD': 'Driver Front Door',
    'ALACARTE_PFD': 'Passenger Front Door',
    'ALACARTE_DRD': 'Driver Rear Door',
    'ALACARTE_PRD': 'Passenger Rear Door',
    'ALACARTE_QTR': 'Quarter Window',
    'ALACARTE_RW': 'Rear Window',
}

# Tokens that may appear in the Services column (col 5) signaling additional
# line items that aren't covered by the dedicated columns 6-11. We pick these
# up by regex-scanning the Services text — they're separated by " | ".
EXTRA_SERVICE_TOKENS = [
    'REMOVAL_SUNSTRIP', 'REMOVAL_2FD', 'REMOVAL_WS', 'REMOVAL_REAR',
    'ALACARTE_DFD', 'ALACARTE_PFD', 'ALACARTE_DRD', 'ALACARTE_PRD',
    'ALACARTE_QTR', 'ALACARTE_RW',
]


def parse_row(row, header):
    """Parse a 2026FWT row dict into a structured form."""
    return {
        'date':            row[0],
        'name':            safe_str(row[1]),
        'phone':           normalize_phone(row[2]),
        'email':           normalize_email(row[3]),
        'vehicle_text':    safe_str(row[4]),
        'services_text':   safe_str(row[5]),
        'full':            safe_str(row[6]),
        'fd2':             safe_str(row[7]),
        'ws':              safe_str(row[8]),
        'ts':              safe_str(row[9]),
        'sr':              safe_str(row[10]),
        'removal':         safe_str(row[11]),
        'subtotal_sheet':  to_num(row[12]),
        'discount_sheet':  to_num(row[13]),
        'discount_note':   safe_str(row[14]),
        'nfw':             to_num(row[15]),
        'tip':             to_num(row[16]),
        'total':           to_num(row[17]),
        'deposit':         to_num(row[18]),
        'balance_due':     to_num(row[19]),
        'payment_method':  safe_str(row[20]),
        'processor':       safe_str(row[21]),
        'source':          safe_str(row[22]),
        'gc':              safe_str(row[23]),
        'code':            safe_str(row[24]),
        'note':            safe_str(row[25]),
        'company':         safe_str(row[26]),
        'starting_total':  to_num(row[27]),
        'invoice_num':     safe_str(row[29]),
        'event_id':        safe_str(row[30]),
    }


def parse_vehicle_text(text):
    """Parse '2015 Optima' or '2009 | Frontier |' → (year, make, model). Make may be missing."""
    if not text:
        return None, None, None
    s = text.replace('|', ' ').strip()
    s = re.sub(r'\s+', ' ', s)
    m = re.match(r'(\d{4})\s+(.+)', s)
    if not m:
        return None, None, None
    year = int(m.group(1))
    rest = m.group(2).strip()
    parts = rest.split()
    if len(parts) == 1:
        return year, None, parts[0]
    return year, parts[0], ' '.join(parts[1:])


def find_nfw_abbrev(services_text):
    """Search for NF abbrevs in the Services column. Returns abbrev or None."""
    if not services_text:
        return None
    upper = services_text.upper()
    # Order matters: longer first
    for abbrev in ('NF2FWS', 'NFFWS', 'NF2F', 'NFWS', 'NFF'):
        if re.search(r'\b' + abbrev + r'\b', upper):
            return abbrev
    return None


# ============================================================================
# BUILD LINE ITEMS WITH MENU PRICES
# ============================================================================
def build_line_items(parsed, class_keys_list):
    """
    Returns (line_items, warnings).

    class_keys_list is the ORDERED list of vehicle class_keys from
    auto_vehicles (e.g. ["FRONT2", "CREW_CAB"]). For each service the lookup
    iterates the list and uses the first class that has a matching pricing row,
    then falls back to '*'. Mirrors src/app/components/booking/pricing.ts.

    Add fee surcharges (ADD_FEE_25/50/75/100) come from get_add_fee() and are
    added on top of the FULL_SIDES line as part of its unit_price.
    """
    items = []
    warnings = []
    sort = 0

    def add_item(service_key, raw_film_shade, force_film=None, force_class=None):
        nonlocal sort
        film_abbrev, shade, shade_front, shade_rear = parse_film_shade(raw_film_shade, allow_match=True)
        film_abbrev = force_film or film_abbrev
        if not film_abbrev:
            warnings.append(f"No film abbrev parsed for {service_key} from {raw_film_shade!r}")
            return None
        film = FILM_BY_ABBREV.get(film_abbrev)
        if not film:
            warnings.append(f"Unknown film abbrev {film_abbrev!r} for {service_key}")
            return None

        # Build the candidate class list for the lookup:
        #   - force_class overrides everything (used for FULL_WS / SUN_STRIP / sunroofs which are universal '*')
        #   - TWO_FRONT_DOORS only looks at FRONT2 / FRONT2_Q
        #   - everything else uses the vehicle's full class_keys list
        if force_class:
            candidates = [force_class]
        elif service_key == 'TWO_FRONT_DOORS':
            candidates = [c for c in class_keys_list if c in ('FRONT2', 'FRONT2_Q')] or ['FRONT2']
        else:
            candidates = list(class_keys_list)

        price, duration, matched_class = lookup_price(candidates, service_key, film['id'])
        if price is None:
            warnings.append(
                f"No auto_pricing for classes={candidates} service={service_key} film_id={film['id']}"
            )
            price = 0.0

        label = LABEL_MAP.get(service_key, service_key)
        custom_fields = {
            'serviceKey': service_key,
            'filmId': film['id'],
            'filmName': film['name'],
            'filmAbbrev': film['abbrev'],
        }
        if shade:
            custom_fields['shade'] = shade
        if shade_front:
            custom_fields['shadeFront'] = shade_front
        if shade_rear:
            custom_fields['shadeRear'] = shade_rear

        item = {
            'description': label,
            'unit_price': round2(price),
            'line_total': round2(price),
            'sort_order': sort,
            'custom_fields': custom_fields,
            '_service_key': service_key,
        }
        items.append(item)
        sort += 1
        return item

    full_item = None
    if parsed['full']:
        full_item = add_item('FULL_SIDES', parsed['full'])
    if parsed['fd2']:
        add_item('TWO_FRONT_DOORS', parsed['fd2'])
    if parsed['ws']:
        add_item('FULL_WS', parsed['ws'], force_class='*')
    if parsed['ts']:
        # Sun strip is always BLK 5%, flat $60
        add_item('SUN_STRIP', parsed['ts'] or 'BLK 5%', force_film='BLK', force_class='*')
    if parsed['sr']:
        sk = 'SUNROOF_PANO' if is_panoramic(parsed['sr']) else 'SUNROOF_SINGLE'
        add_item(sk, parsed['sr'], force_class='*')
    if parsed['removal']:
        # Removal column may be a film label or a dollar amount
        rem_text = parsed['removal']
        rem_amt = None
        try:
            rem_amt = float(rem_text)
        except (TypeError, ValueError):
            pass
        if rem_amt is not None and rem_amt > 0:
            items.append({
                'description': LABEL_MAP['REMOVAL_FULL'],
                'unit_price': round2(rem_amt),
                'line_total': round2(rem_amt),
                'sort_order': sort,
                'custom_fields': {'serviceKey': 'REMOVAL_FULL'},
                '_service_key': 'REMOVAL_FULL',
            })
            sort += 1
        else:
            price, _, _ = lookup_price(class_keys_list, 'REMOVAL_FULL', None)
            if price is None:
                warnings.append(f"No REMOVAL_FULL price for classes={class_keys_list}")
                price = 0.0
            items.append({
                'description': LABEL_MAP['REMOVAL_FULL'],
                'unit_price': round2(price),
                'line_total': round2(price),
                'sort_order': sort,
                'custom_fields': {'serviceKey': 'REMOVAL_FULL'},
                '_service_key': 'REMOVAL_FULL',
            })
            sort += 1

    # NOTE: We deliberately do NOT auto-apply ADD_FEE_* surcharges here, even
    # though the live booking flow does. Reason: historic spreadsheet rows often
    # show the BASE price was charged with no add fee — meaning the team waived
    # the fee at the counter, or the vehicle wasn't classified with the add fee
    # yet at the time of service. The Manual Adjustment fallback below detects
    # when the spreadsheet target sub IS higher than the line items by the exact
    # add_fee amount, and only then adds it as an explicit "Add Fee" line.

    # Scan the Services text column for extra service tokens that don't have
    # dedicated columns (REMOVAL_SUNSTRIP, ALACARTE_*, etc.). These appear in
    # rows that have multiple removals or à-la-carte additions.
    services_text = parsed.get('services_text') or ''
    for token in EXTRA_SERVICE_TOKENS:
        if not re.search(r'\b' + re.escape(token) + r'\b', services_text):
            continue
        # Look up the price (universal '*' first, then class-specific)
        price, _, _ = lookup_price(class_keys_list, token, None)
        if price is None:
            # Try film-keyed lookup with BLK as default for removals
            price, _, _ = lookup_price(class_keys_list, token, 5)
        if price is None:
            warnings.append(f"No price for extra service {token}")
            price = 0.0
        items.append({
            'description': LABEL_MAP.get(token, token),
            'unit_price': round2(price),
            'line_total': round2(price),
            'sort_order': sort,
            'custom_fields': {'serviceKey': token},
            '_service_key': token,
        })
        sort += 1

    return items, warnings


# ============================================================================
# BUILD APPLIED DISCOUNTS
# ============================================================================
def parse_discount_note(note):
    """Returns a list of (kind, dollar_amt) inferred from the note text."""
    if not note:
        return []
    out = []
    upper = note.upper()
    if 'CASH' in upper and '5%' in upper:
        out.append('cash_5')
    if 'MILITARY' in upper:
        # Try to extract dollar amount
        m = re.search(r'\$?\s*(\d+)', note)
        if m:
            out.append(('military', int(m.group(1))))
        else:
            out.append('military_25')
    return out


def build_applied_discounts(parsed, line_items_sum):
    """
    Returns (applied_discounts_jsonb, total_discount_amount).

    Stacking order (Option B — matches the real-world counter checkout):
      1. Friend/Family / custom discount (off the menu subtotal)
      2. Military dollar discount(s) (off whatever's left)
      3. Cash 5% (off whatever's left after the above)

    Anchoring strategy:
      - Trust the spreadsheet's "Discount" column (col 13) as the SUM of the
        OFFICIAL discounts (military + cash). Distribute it across the named
        types listed in the discount note: military takes its flat dollar,
        cash takes the rest.
      - Compute the friend/custom discount as the leftover needed to land on
        the spreadsheet total: friend = line_items_sum + nfw + tip - parsed_total - col13
      - This guarantees the recorded discounts always tie out to the spreadsheet
        and to the actual cash collected.
    """
    parsed_total = round2(parsed['total'])
    nfw = round2(parsed['nfw'])
    tip = round2(parsed['tip'])
    sheet_discount = round2(parsed['discount_sheet'])

    # Total discount that has to come off line items to land at parsed_total
    total_discount_needed = round2(line_items_sum + nfw + tip - parsed_total)
    if total_discount_needed < -0.005:
        # parsed total > menu — shouldn't happen, but bail safely
        return [], 0.0

    note_parts = parse_discount_note(parsed['discount_note'])
    has_cash = 'cash_5' in note_parts
    military_amts = [p[1] for p in note_parts if isinstance(p, tuple) and p[0] == 'military']
    military_total = round2(sum(military_amts))

    # Cash portion = sheet_discount - military_total (what's left for cash inside col 13)
    cash_amt = round2(sheet_discount - military_total) if has_cash else 0.0
    if cash_amt < 0:
        cash_amt = 0.0

    # Unattributed sheet discount = sheet_discount minus everything we recognized
    # in the note. Happens when the spreadsheet has a discount but the note text
    # doesn't say "Cash 5%" or "Military" — usually a GC redemption, manager
    # comp, or a one-off manual discount with no note. Record it as a Manual
    # Discount entry rather than silently losing it.
    attributed = military_total + cash_amt
    manual_amt = round2(sheet_discount - attributed)
    if manual_amt < 0.005:
        manual_amt = 0.0

    # Friend = additional discount needed beyond what the spreadsheet records.
    # Used when the customer got a stacked discount (cash/military + a friend
    # favor) where only one half made it onto the spreadsheet's discount column.
    friend_amt = round2(total_discount_needed - sheet_discount)
    if friend_amt < 0.005:
        friend_amt = 0.0

    discounts = []

    if friend_amt > 0:
        discounts.append({
            "discount_type_id": None,
            "name": "Friend/Family Discount",
            "discount_type": "dollar",
            "discount_value": friend_amt,
            "applies_to": "subtotal",
            "include_warranty": False,
            "amount": friend_amt,
        })

    if manual_amt > 0:
        discounts.append({
            "discount_type_id": None,
            "name": "Manual Discount",
            "discount_type": "dollar",
            "discount_value": manual_amt,
            "applies_to": "subtotal",
            "include_warranty": False,
            "amount": manual_amt,
        })

    for amt in military_amts:
        disc_type = MILITARY_25_DISCOUNT_TYPE if amt >= 25 else MILITARY_10_DISCOUNT_TYPE
        entry = dict(disc_type)
        entry['discount_value'] = float(amt)
        entry['amount'] = round2(amt)
        discounts.append(entry)

    if has_cash and cash_amt > 0:
        entry = dict(CASH_DISCOUNT_TYPE)
        entry['amount'] = cash_amt
        discounts.append(entry)

    final_total = round2(sum(d['amount'] for d in discounts))
    return discounts, final_total


# ============================================================================
# BUILD ALL RECORDS FOR ONE ROW
# ============================================================================
def build_records_for_row(parsed, bl_match):
    """
    Returns a dict with the proposed booking, document, line_items, payment,
    plus warnings + math summary.
    """
    warnings = []
    # Vehicle / class lookup — always look up auto_vehicles by year/make/model
    # (the source of truth for class_keys), then merge in any ADD_FEE_* entries
    # that the BookingLog row had appended manually.
    if bl_match and bl_match.get('vehicle_make'):
        vyear = bl_match.get('vehicle_year')
        vmake = bl_match.get('vehicle_make')
        vmodel = bl_match.get('vehicle_model')
    else:
        vyear, vmake, vmodel = parse_vehicle_text(parsed['vehicle_text'])

    _, class_keys_list = lookup_vehicle(vyear, vmake, vmodel)
    if not class_keys_list:
        warnings.append(
            f"Could not resolve class for {vyear} {vmake} {vmodel} from text {parsed['vehicle_text']!r}"
        )
        class_keys_list = []

    # Merge any ADD_FEE_* entries from the BookingLog class_keys string
    # (BookingLog can manually annotate fees that aren't in the auto_vehicles row)
    if bl_match and bl_match.get('class_keys'):
        bl_extras = [c.strip() for c in re.split(r'[|,]', bl_match['class_keys']) if c.strip()]
        for extra in bl_extras:
            if extra.startswith('ADD_FEE_') and extra not in class_keys_list:
                class_keys_list = list(class_keys_list) + [extra]

    # The first non-FRONT class is the canonical "vehicle class" stored on the
    # booking/document records (used for display purposes only)
    primary_class = next(
        (c for c in class_keys_list if not c.startswith('FRONT') and not c.startswith('ADD_FEE')),
        class_keys_list[0] if class_keys_list else None
    )

    # Build line items at full menu price
    line_items, lw = build_line_items(parsed, class_keys_list)
    warnings.extend(lw)

    line_items_sum = round2(sum(li['line_total'] for li in line_items))

    # If the line items don't add up to what the spreadsheet implies, add a
    # "Manual Adjustment" line for the remainder. The target subtotal is what
    # the line items need to total to so that
    #   subtotal - sheet_discount + nfw + tip = parsed_total
    # → target = parsed_total + sheet_discount - nfw - tip
    # This catches: missing services (Shiva Dayani's REMOVAL_SUNSTRIP wasn't
    # handled), unknown alacarte items (Sean Dowdie's ALACARTE_DFD has no
    # menu price), and price drift where the menu has changed since April.
    parsed_total_for_calc = round2(parsed['total'])
    nfw_for_calc = round2(parsed['nfw'])
    tip_for_calc = round2(parsed['tip'])
    sheet_disc_for_calc = round2(parsed['discount_sheet'])
    target_subtotal = round2(parsed_total_for_calc + sheet_disc_for_calc - nfw_for_calc - tip_for_calc)
    gap = round2(target_subtotal - line_items_sum)
    if gap > 0.5:
        # Smart label: if the gap exactly matches an ADD_FEE_* amount on this
        # vehicle, the team did charge the fee that day — label the line as
        # "Add Fee" rather than the generic "Manual Adjustment". Otherwise it's
        # menu drift / unknown service / etc.
        expected_add_fee = get_add_fee(class_keys_list)
        if expected_add_fee > 0 and abs(gap - expected_add_fee) < 0.5:
            label = 'Add Fee'
            svc_key = 'ADD_FEE'
        else:
            label = 'Manual Adjustment'
            svc_key = 'MANUAL_ADJUSTMENT'
        line_items.append({
            'description': label,
            'unit_price': gap,
            'line_total': gap,
            'sort_order': len(line_items),
            'custom_fields': {'serviceKey': svc_key, 'historicAdjustment': True},
            '_service_key': svc_key,
        })
        line_items_sum = round2(line_items_sum + gap)
        warnings.append(f"Added ${gap:.2f} {label} line to close gap")

    # Build discounts (this includes the friend variance if any)
    applied_discounts, total_discount = build_applied_discounts(parsed, line_items_sum)
    discount_note = parsed['discount_note']

    # Warranty
    nfw_abbrev = find_nfw_abbrev(parsed['services_text'])
    applied_warranty = None
    nfw_amount_actual = round2(parsed['nfw'])
    if nfw_abbrev:
        opt = NFW_BY_ABBREV[nfw_abbrev]
        applied_warranty = {
            "warranty_product_id": opt['warranty_product_id'],
            "option_id": opt['option_id'],
            "name": opt['name'],
            "option_name": opt['option_name'],
            "price": opt['price'],
        }
        if abs(opt['price'] - nfw_amount_actual) > 0.5:
            warnings.append(
                f"NFW abbrev {nfw_abbrev} expects ${opt['price']} but spreadsheet shows ${nfw_amount_actual}"
            )

    # Math sanity: line_items_sum + nfw + tip - total_discount should == parsed['total']
    expected_total = round2(line_items_sum + nfw_amount_actual + parsed['tip'] - total_discount)
    parsed_total = round2(parsed['total'])
    math_ok = abs(expected_total - parsed_total) < 0.02
    if not math_ok:
        warnings.append(
            f"Math mismatch: line_sum({line_items_sum}) + nfw({nfw_amount_actual}) + tip({parsed['tip']}) "
            f"- discount({total_discount}) = {expected_total}, but spreadsheet total = {parsed_total}"
        )

    # Original booking total (frozen via starting_total_override)
    if bl_match and bl_match.get('subtotal', 0) > 0:
        starting_total = round2(bl_match['subtotal'])
    elif parsed['starting_total'] > 0:
        starting_total = round2(parsed['starting_total'])
    else:
        starting_total = line_items_sum

    upsell_amount = max(0.0, round2(line_items_sum - starting_total))

    # Document subtotal = full menu sum (matches line item totals)
    doc_subtotal = line_items_sum
    deposit = round2(parsed['deposit'])

    # Build the services_json snapshot in the live TintServiceLine shape
    # (camelCase fields the dashboard's AppointmentCard + EditAppointmentModal expect).
    # This represents the FINAL services as actually performed — what the timeline
    # card should display. The frozen original total is preserved separately via
    # starting_total_override so upsell tracking still works.
    services_json_live = []
    for li in line_items:
        cf = li['custom_fields']
        svc_entry = {
            'serviceKey': cf.get('serviceKey'),
            'label': li['description'],
            'filmId': cf.get('filmId'),
            'filmName': cf.get('filmName'),
            'filmAbbrev': cf.get('filmAbbrev'),
            'shadeFront': cf.get('shadeFront'),
            'shadeRear': cf.get('shadeRear'),
            'shade': cf.get('shade'),
            'price': li['unit_price'],
            'discountAmount': 0,
            'duration': 0,
        }
        services_json_live.append(svc_entry)

    # Booking date / time / type — prefer BookingLog truth
    if bl_match:
        appt_date = bl_match['date'].date() if hasattr(bl_match['date'], 'date') else bl_match['date']
        appt_time = bl_match['time']
        duration = bl_match['duration']
        appt_type_raw = (bl_match.get('appointment_type') or '').lower()
        if appt_type_raw in ('dropoff', 'waiting', 'headsup_30', 'headsup_60'):
            appt_type = appt_type_raw
        else:
            appt_type = 'dropoff'
        raw_src = (bl_match.get('booking_source') or '').lower()
        if raw_src in ('nodep', 'shopify', 'online'):
            booking_source = 'online'
        elif raw_src == 'phone':
            booking_source = 'phone'
        elif raw_src in ('gc', 'gift_certificate'):
            booking_source = 'gc'
        elif raw_src == 'sameday':
            booking_source = 'sameday'
        else:
            booking_source = 'internal'
    else:
        appt_date = parsed['date'].date() if hasattr(parsed['date'], 'date') else parsed['date']
        appt_time = None
        duration = None
        appt_type = 'dropoff'
        booking_source = 'internal'

    # Build the records
    booking = {
        'shop_id': SHOP_ID,
        'booking_id': parsed['invoice_num'],  # match prior import for stable IDs
        'booking_source': booking_source,
        'appointment_type': appt_type,
        'service_type': 'tint',
        'emoji_marker': bl_match.get('emoji_marker') if bl_match else None,
        'customer_name': (bl_match.get('name') if bl_match else None) or parsed['name'],
        'customer_phone': parsed['phone'] or (normalize_phone(bl_match.get('phone')) if bl_match else None),
        'customer_email': parsed['email'] or (normalize_email(bl_match.get('email')) if bl_match else None),
        'vehicle_year': vyear,
        'vehicle_make': vmake,
        'vehicle_model': vmodel,
        'class_keys': '|'.join(class_keys_list) if class_keys_list else None,
        'appointment_date': to_date_str(appt_date),
        'appointment_time': to_time_str(appt_time),
        'duration_minutes': duration,
        # FINAL services in live TintServiceLine shape — what the timeline card renders
        'services_json': services_json_live,
        # Subtotal reflects the final services (matches line items / edit modal sum)
        'subtotal': doc_subtotal,
        # Frozen original — preserves upsell math regardless of edits to subtotal
        'starting_total_override': starting_total,
        'discount_code': bl_match.get('discount_code') if bl_match else None,
        'discount_amount': bl_match.get('discount_amount', 0) if bl_match else 0,
        'deposit_paid': bl_match.get('deposit_paid', 0) if bl_match else deposit,
        'balance_due': 0,
        'total_paid': parsed_total,
        'payment_method': (parsed['payment_method'] or 'cash').lower().replace(' ', '_'),
        'calendar_event_id': parsed['event_id'],
        # 'invoiced' is the status the live counter checkout writes after payment
        # (see src/app/api/documents/[id]/payments/route.ts:71). The dashboard's
        # TimelineView treats invoiced/paid as "isPaid" and turns the left block
        # green. 'completed' would render as not-yet-paid.
        'status': 'invoiced',
        'module': 'auto_tint',
        'import_source': 'historic_import',
        'notes': parsed['discount_note'],
    }

    document = {
        'shop_id': SHOP_ID,
        'doc_number': f"INV-{parsed['invoice_num']}",
        'doc_type': 'invoice',
        'checkout_type': 'counter',
        'status': 'paid',
        'customer_name': parsed['name'],
        'customer_phone': parsed['phone'],
        'customer_email': parsed['email'],
        'vehicle_year': vyear,
        'vehicle_make': vmake,
        'vehicle_model': vmodel,
        'class_keys': '|'.join(class_keys_list) if class_keys_list else None,
        'subtotal': doc_subtotal,
        'starting_total': starting_total,
        'upsell_amount': upsell_amount,
        'discount_amount': total_discount,
        'discount_note': discount_note,
        'applied_discounts': applied_discounts,
        'applied_warranty': applied_warranty,
        'deposit_paid': deposit,
        'balance_due': 0,
        'total_paid': parsed_total,
        'tip_amount': round2(parsed['tip']),
        'payment_method': (parsed['payment_method'] or 'cash').lower().replace(' ', '_'),
        'payment_processor': (parsed['processor'] or 'manual').lower(),
        'tax_rate': 0,
        'tax_amount': 0,
        'paid_at': to_local_noon_iso(parsed['date']),
        'created_at': to_local_noon_iso(parsed['date']),
        'import_source': 'historic_import',
    }

    # Strip the helper _service_key field before insert
    line_items_clean = [
        {k: v for k, v in li.items() if not k.startswith('_')}
        for li in line_items
    ]
    for li in line_items_clean:
        li['module'] = 'auto_tint'
        li['quantity'] = 1

    payment = {
        'shop_id': SHOP_ID,
        'amount': parsed_total,
        'payment_method': (parsed['payment_method'] or 'cash').lower().replace(' ', '_'),
        'processor': (parsed['processor'] or 'manual').lower(),
        'status': 'confirmed',
        'created_at': to_local_noon_iso(parsed['date']),
    }

    return {
        'booking': booking,
        'document': document,
        'line_items': line_items_clean,
        'payment': payment,
        'warnings': warnings,
        'math': {
            'line_items_sum': line_items_sum,
            'nfw': nfw_amount_actual,
            'tip': round2(parsed['tip']),
            'total_discount': total_discount,
            'expected_total': expected_total,
            'spreadsheet_total': parsed_total,
            'spreadsheet_subtotal': round2(parsed['subtotal_sheet']),
            'spreadsheet_discount': round2(parsed['discount_sheet']),
            'starting_total': starting_total,
            'upsell_amount': upsell_amount,
        }
    }


# ============================================================================
# LOAD ROWS
# ============================================================================
def load_april_rows():
    wb = openpyxl.load_workbook(FINANCIAL_XLSX, data_only=True)
    ws = wb['2026FWT']
    header = [c.value for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        d = row[0]
        if not d or not hasattr(d, 'year') or d.year != 2026 or d.month != 4:
            continue
        if not safe_str(row[29]):
            continue
        out.append((r, parse_row(row, header)))
    return out


def find_row_by_invoice(invoice_num):
    for r_idx, parsed in load_april_rows():
        if parsed['invoice_num'] == invoice_num:
            return r_idx, parsed
    return None, None


# ============================================================================
# DISPLAY
# ============================================================================
def show_records(records):
    p = records
    print()
    print("=" * 78)
    print(f"PROPOSED IMPORT FOR {p['document']['doc_number']}")
    print("=" * 78)
    print()
    print("BOOKING (auto_bookings):")
    for k, v in p['booking'].items():
        if k == 'services_json' and v:
            print(f"  {k}:")
            for svc in v:
                print(f"    - {json.dumps(svc, default=str)}")
        else:
            print(f"  {k}: {v!r}")
    print()
    print("DOCUMENT (documents):")
    for k, v in p['document'].items():
        if k in ('applied_discounts', 'applied_warranty') and v:
            print(f"  {k}:")
            print(f"    {json.dumps(v, indent=4, default=str)}")
        else:
            print(f"  {k}: {v!r}")
    print()
    print(f"LINE ITEMS ({len(p['line_items'])}):")
    for li in p['line_items']:
        print(f"  - sort {li['sort_order']}: {li['description']!r}")
        print(f"      unit_price=${li['unit_price']} line_total=${li['line_total']}")
        print(f"      custom_fields={li['custom_fields']}")
    print()
    print("PAYMENT (document_payments):")
    for k, v in p['payment'].items():
        print(f"  {k}: {v!r}")
    print()
    print("MATH SUMMARY:")
    m = p['math']
    print(f"  line items sum (full menu):    ${m['line_items_sum']:>10.2f}")
    print(f"  + NFW:                         ${m['nfw']:>10.2f}")
    print(f"  + tip:                         ${m['tip']:>10.2f}")
    print(f"  - total discount:              ${m['total_discount']:>10.2f}")
    print(f"  = expected total:              ${m['expected_total']:>10.2f}")
    print(f"  spreadsheet total:             ${m['spreadsheet_total']:>10.2f}")
    print(f"  match? {'YES' if abs(m['expected_total'] - m['spreadsheet_total']) < 0.02 else 'NO'}")
    print()
    print(f"  starting_total (original):     ${m['starting_total']:>10.2f}")
    print(f"  upsell (line_sum - starting):  ${m['upsell_amount']:>10.2f}")
    print()
    print(f"  spreadsheet sub column:        ${m['spreadsheet_subtotal']:>10.2f}")
    print(f"  spreadsheet discount column:   ${m['spreadsheet_discount']:>10.2f}")
    print()
    if p['warnings']:
        print("WARNINGS:")
        for w in p['warnings']:
            print(f"  ! {w}")
    else:
        print("WARNINGS: none")
    print()


# ============================================================================
# WIPE
# ============================================================================
def wipe_april():
    """
    Delete all April 2026 historic_import documents AND auto_bookings.

    Identifies the April rows in two passes:
      1. Documents in April via import_source + created_at window
      2. auto_bookings with booking_id matching the April spreadsheet invoice
         numbers (catches orphans where document_id was unlinked but the row
         was never deleted, e.g. from a failed mid-import)
    Cascades line items, payments, document records, and finally bookings.
    """
    print("=== WIPING APRIL 2026 historic_import data ===")

    # Pass 1 — find all April docs by created_at window
    docs = sb("GET", "documents", params={
        "select": "id,doc_number",
        "shop_id": f"eq.{SHOP_ID}",
        "import_source": "eq.historic_import",
        "created_at": "gte.2026-04-01",
        "and": "(created_at.lt.2026-05-01)",
    })
    print(f"Found {len(docs)} April documents to wipe")

    if docs:
        doc_ids = [d['id'] for d in docs]
        # Delete payments
        print("  Deleting document_payments...")
        sb("DELETE", "document_payments", params={
            "document_id": f"in.({','.join(doc_ids)})"
        })
        # Delete line items
        print("  Deleting document_line_items...")
        sb("DELETE", "document_line_items", params={
            "document_id": f"in.({','.join(doc_ids)})"
        })
        # Null out auto_bookings.document_id so the FK doesn't block doc deletion
        print("  Unlinking auto_bookings.document_id...")
        sb("PATCH", "auto_bookings",
           body={"document_id": None},
           params={"document_id": f"in.({','.join(doc_ids)})"})
        # Delete documents
        print("  Deleting documents...")
        sb("DELETE", "documents", params={
            "id": f"in.({','.join(doc_ids)})"
        })

    # Pass 2 — find all April-numbered auto_bookings (catches orphans).
    # Pre-load the canonical April invoice list from the spreadsheet so we
    # only delete rows whose booking_id is unambiguously a 2026-04 row.
    april_rows = load_april_rows()
    april_invoice_nums = sorted({p['invoice_num'] for _, p in april_rows if p['invoice_num']})
    if april_invoice_nums:
        # PostgREST 'in' filter quoting: wrap each value in double quotes if it
        # contains a hyphen, comma, or space. The values here are like '260401-001'.
        in_list = ','.join(f'"{v}"' for v in april_invoice_nums)
        bookings = sb("GET", "auto_bookings", params={
            "select": "id,booking_id,customer_name,document_id",
            "shop_id": f"eq.{SHOP_ID}",
            "import_source": "eq.historic_import",
            "booking_id": f"in.({in_list})",
        }) or []
        if bookings:
            print(f"Found {len(bookings)} April-numbered auto_bookings to wipe")
            booking_uuids = [b['id'] for b in bookings]
            print(f"  Deleting {len(booking_uuids)} auto_bookings...")
            sb("DELETE", "auto_bookings", params={
                "id": f"in.({','.join(booking_uuids)})"
            })

    print("=== WIPE COMPLETE ===")


# ============================================================================
# INSERT
# ============================================================================
def insert_records(records):
    booking = records['booking']
    document = records['document']
    line_items = records['line_items']
    payment = records['payment']

    print(f"  Inserting auto_bookings {booking['booking_id']}...")
    booking_resp = sb("POST", "auto_bookings", body=booking)
    booking_id = booking_resp[0]['id']

    # Link the document FK to the booking before insert (bidirectional link)
    document['booking_id'] = booking_id

    print(f"  Inserting documents {document['doc_number']}...")
    doc_resp = sb("POST", "documents", body=document)
    doc_id = doc_resp[0]['id']

    if line_items:
        print(f"  Inserting {len(line_items)} document_line_items...")
        for li in line_items:
            li['document_id'] = doc_id
        sb("POST", "document_line_items", body=line_items)

    print(f"  Inserting document_payments...")
    payment['document_id'] = doc_id
    sb("POST", "document_payments", body=payment)

    print(f"  Linking auto_bookings.document_id...")
    sb("PATCH", "auto_bookings",
       body={"document_id": doc_id},
       params={"id": f"eq.{booking_id}"})

    print(f"  → booking={booking_id}, doc={doc_id}")
    return booking_id, doc_id


# ============================================================================
# MAIN
# ============================================================================
def main():
    load_env()
    args = sys.argv[1:]

    if "--show" in args:
        invoice = args[args.index("--show") + 1]
        load_booking_log()
        r_idx, parsed = find_row_by_invoice(invoice)
        if not parsed:
            print(f"No row found for invoice {invoice}")
            return
        bl_match = _booking_log_by_event.get(parsed['event_id']) if parsed['event_id'] else None
        records = build_records_for_row(parsed, bl_match)
        print(f"Spreadsheet row {r_idx} → {parsed['name']!r} | {parsed['vehicle_text']!r}")
        if bl_match:
            print(f"  BookingLog match: {bl_match['booking_id']} | "
                  f"{bl_match['vehicle_year']} {bl_match['vehicle_make']} {bl_match['vehicle_model']} "
                  f"| class={bl_match['class_keys']} | original ${bl_match['subtotal']}")
        else:
            print("  No BookingLog match (event_id missing or unmatched)")
        show_records(records)
        return

    if "--wipe" in args:
        if "--yes" not in args:
            resp = input("Wipe ALL April 2026 historic_import documents + bookings? [yes/no] ")
            if resp.strip().lower() != 'yes':
                print("Aborted.")
                return
        wipe_april()
        return

    if "--import" in args:
        invoice = args[args.index("--import") + 1]
        load_booking_log()
        r_idx, parsed = find_row_by_invoice(invoice)
        if not parsed:
            print(f"No row found for invoice {invoice}")
            return
        bl_match = _booking_log_by_event.get(parsed['event_id']) if parsed['event_id'] else None
        records = build_records_for_row(parsed, bl_match)
        show_records(records)
        if records['warnings']:
            print("Refusing to import due to warnings. Use --import-force to override.")
            if "--import-force" not in args:
                return
        if "--yes" not in args:
            resp = input(f"Insert this row? [yes/no] ")
            if resp.strip().lower() != 'yes':
                print("Aborted.")
                return
        insert_records(records)
        print("DONE.")
        return

    if "--import-all" in args:
        load_booking_log()
        rows = load_april_rows()
        print(f"Loaded {len(rows)} April rows")
        if "--yes" not in args:
            resp = input(f"Insert all {len(rows)} rows? [yes/no] ")
            if resp.strip().lower() != 'yes':
                print("Aborted.")
                return
        for r_idx, parsed in rows:
            bl_match = _booking_log_by_event.get(parsed['event_id']) if parsed['event_id'] else None
            records = build_records_for_row(parsed, bl_match)
            print(f"\n--- {parsed['invoice_num']} {parsed['name']} ---")
            if records['warnings']:
                for w in records['warnings']:
                    print(f"  ! {w}")
            insert_records(records)
        print("\nALL DONE.")
        return

    print(__doc__)


if __name__ == '__main__':
    main()
