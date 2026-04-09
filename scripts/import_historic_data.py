#!/usr/bin/env python3
"""
Historic Data Import (Session 9, 2026-04-09)
=============================================
Imports:
  - 2024 from FWT Appointments → auto_bookings (metrics-only, no docs)
  - 2025 from FWT Appointments → auto_bookings (metrics-only, no docs)
  - 2026 BookingLog → auto_bookings (originals with starting_total)
  - 2026FWT → documents + line items + payment records

Direct Supabase REST API writes. NO API routes touched. NO SMS/email sent.
Transaction ledger is NOT modified (entries already exist from spreadsheet sync).
"""

import openpyxl
import json
import re
import os
import sys
import urllib.request
import urllib.error
from datetime import datetime, date
from collections import defaultdict

# ============================================================================
# CONFIG
# ============================================================================
SUPABASE_URL = "https://jyrgqbhpdbaiedfgoyse.supabase.co"
SERVICE_KEY = None  # Loaded from .env.local

SHOP_ID = 1
APPOINTMENTS_XLSX = "legacy/spreadsheets/FWT Appointments.xlsx"
BOOKINGLOG_XLSX = "legacy/spreadsheets/FWT_Booking_Spec_FINAL (1).xlsx"
FINANCIAL_XLSX = "legacy/spreadsheets/2026 Financial Core – FWT & FWG (1).xlsx"

# Service key mappings for the 2026FWT columns
SERVICE_KEY_MAP = {
    "FULL": "FULL_SIDES",
    "2FD": "TWO_FRONT_DOORS",
    "WS": "FULL_WS",
    "TS": "SUN_STRIP",
    "SR": "SUNROOF_SINGLE",  # may be PANO; will guess from row text
    "Removal": "REMOVAL_FULL",
}

# ============================================================================
# ENV LOADER
# ============================================================================
def load_env():
    global SERVICE_KEY
    with open(".env.local") as f:
        for line in f:
            if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
                SERVICE_KEY = line.split("=", 1)[1].strip()
                return
    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY not found")


# ============================================================================
# SUPABASE REST CLIENT
# ============================================================================
def supabase_request(method, path, body=None, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        qs = "&".join(f"{k}={v}" for k, v in params.items())
        url = f"{url}?{qs}"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text else None
    except urllib.error.HTTPError as e:
        body_text = e.read().decode()
        # Print first row of body if it's a list to debug
        if isinstance(body, list) and len(body) > 0:
            print(f"  First row in failing batch: {json.dumps(body[0], default=str)[:500]}", file=sys.stderr)
        print(f"HTTP {e.code}: {body_text}", file=sys.stderr)
        raise

def insert_batch(table, rows, batch_size=100):
    """Insert rows in batches to avoid request size limits."""
    inserted = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i+batch_size]
        result = supabase_request("POST", table, body=batch)
        inserted += len(result) if result else 0
        if i % 500 == 0 and i > 0:
            print(f"    inserted {inserted}/{len(rows)}")
    return inserted


# ============================================================================
# HELPERS
# ============================================================================
def normalize_phone(raw):
    if raw is None: return None
    s = str(raw).split('.')[0].strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) == 11 and digits.startswith('1'): digits = digits[1:]
    return digits if len(digits) == 10 else None

def normalize_email(raw):
    if raw is None: return None
    s = str(raw).strip().lower()
    return s if '@' in s and '.' in s else None

def parse_vehicle(s):
    """Parse '2022 Ford Mustang' → (2022, 'Ford', 'Mustang')"""
    if not s: return (None, None, None)
    s = str(s).strip()
    m = re.match(r'(\d{4})\s+(\w+)(?:\s+(.+))?', s)
    if m:
        year = int(m.group(1))
        return (year, m.group(2), m.group(3))
    parts = s.split()
    if len(parts) == 0: return (None, None, None)
    if len(parts) == 1: return (None, parts[0], None)
    return (None, parts[0], ' '.join(parts[1:]))

def to_date_str(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    return str(v)

def to_time_str(v):
    if v is None: return None
    if hasattr(v, 'hour'): return v.strftime('%H:%M:%S')
    return str(v)

def to_num(v, default=0):
    if v is None: return default
    try: return float(v)
    except: return default

def safe_str(v):
    return None if v is None else str(v).strip()


# ============================================================================
# IMPORT 2024 / 2025 from FWT Appointments
# ============================================================================
def import_yearly_sheet(year):
    print(f"\n=== Importing {year} from FWT Appointments ===")
    wb = openpyxl.load_workbook(APPOINTMENTS_XLSX, data_only=True)
    ws = wb[str(year)]

    rows_to_insert = []
    skipped = 0

    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        # Columns: Date(2), Name(3), Phone(4), Email(5), Appointment(6), Vehicle(7),
        # Shade(8), WS(9), TS(10), SR(11), Removal(12), Price(13), WFI(14), Tip(15),
        # TOTAL(16), Payment(17), Source(18), Note(19), Company(20), GC(21),
        # Invoice#(22), Record ID(23), Event ID(24)

        d = row[1]
        if not d or not hasattr(d, 'year') or d.year != year:
            skipped += 1
            continue

        name = safe_str(row[2])
        if not name:
            skipped += 1
            continue

        # Build services array based on which service columns are filled
        services = []
        shade = safe_str(row[7])
        ws_val = safe_str(row[8])
        ts_val = safe_str(row[9])
        sr_val = safe_str(row[10])
        rem_val = safe_str(row[11])
        price = to_num(row[12])
        total = to_num(row[15])
        appointment_text = safe_str(row[5])

        # Primary service: infer from appointment text (FULL vs 2FD)
        upper = (appointment_text or '').upper()
        if 'FULL' in upper or 'R&R' in upper:
            services.append({
                'service_key': 'FULL_SIDES',
                'label': 'Full (Sides & Rear)',
                'shade': shade,
                'price': price,
            })
        elif '2FD' in upper or 'PT' in upper:
            services.append({
                'service_key': 'TWO_FRONT_DOORS',
                'label': '2 Front Doors',
                'shade': shade,
                'price': price,
            })
        elif rem_val and not (ws_val or ts_val or sr_val) and price == 0:
            # Removal-only job
            services.append({
                'service_key': 'REMOVAL_FULL',
                'label': 'Removal',
                'price': total,
            })

        if ws_val:
            services.append({'service_key': 'FULL_WS', 'label': 'Full Windshield', 'shade': ws_val})
        if ts_val:
            services.append({'service_key': 'SUN_STRIP', 'label': 'Sun Strip', 'shade': ts_val})
        if sr_val:
            services.append({'service_key': 'SUNROOF_SINGLE', 'label': 'Sunroof', 'shade': sr_val})
        if rem_val and (ws_val or ts_val or sr_val or 'FULL' in upper or '2FD' in upper):
            services.append({'service_key': 'REMOVAL_FULL', 'label': 'Removal'})

        if not services:
            services = [{'service_key': 'OTHER', 'label': appointment_text or 'Service', 'price': total}]

        vyear, vmake, vmodel = parse_vehicle(safe_str(row[6]))
        # Build a unique booking_id: year + row number guarantees uniqueness
        invoice_num = f"H{year}-{r:05d}"

        booking = {
            'shop_id': SHOP_ID,
            'booking_id': invoice_num,
            'booking_source': 'internal',
            'appointment_type': 'dropoff',
            'service_type': 'tint',
            'customer_name': name,
            'customer_phone': normalize_phone(row[3]),
            'customer_email': normalize_email(row[4]),
            'vehicle_year': vyear,
            'vehicle_make': vmake,
            'vehicle_model': vmodel,
            'appointment_date': d.date().isoformat(),
            'services_json': services,
            'subtotal': total,  # Final total for historic; no upsell tracking
            'balance_due': 0,
            'total_paid': total,
            'payment_method': safe_str(row[16]),
            'status': 'completed',
            'module': 'auto_tint',
            'import_source': 'historic_import',
            'notes': safe_str(row[18]),
            'calendar_event_id': safe_str(row[23]),
        }
        rows_to_insert.append(booking)

    print(f"  Prepared {len(rows_to_insert)} rows ({skipped} skipped)")
    inserted = insert_batch('auto_bookings', rows_to_insert, batch_size=200)
    print(f"  Inserted {inserted} bookings for {year}")
    return inserted


# ============================================================================
# IMPORT 2026 BookingLog → auto_bookings (originals)
# ============================================================================
def import_bookinglog_2026():
    print("\n=== Importing 2026 BookingLog (originals) ===")
    wb = openpyxl.load_workbook(BOOKINGLOG_XLSX, data_only=True)
    ws = wb['BookingLog']

    rows_to_insert = []
    skipped_future = 0
    skipped_other = 0

    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        # cols: 1=booking_id, 2=created_at, 3=booking_source, 4=service_type, 5=appt_type,
        #       6=emoji, 7=name, 8=email, 9=phone, 10=year, 11=make, 12=model, 13=class_keys,
        #       14=appt_date, 15=appt_time, 16=duration, 17=services_json, 18=cal_title,
        #       19=subtotal, 20=disc_code, 21=disc_type, 22=disc_percent, 23=disc_amt,
        #       24=deposit, 25=balance, 26=cal_event_id, 27=shopify, 28=status,
        #       29=add_interests, 30=multi_grp, 31=notes, 32=window_status

        booking_id = safe_str(row[0])
        if not booking_id:
            skipped_other += 1
            continue

        d = row[13]
        if not d or not hasattr(d, 'year') or d.year != 2026:
            skipped_other += 1
            continue

        # Skip future appointments per Joe's instructions
        if d > datetime(2026, 4, 9):
            skipped_future += 1
            continue

        name = safe_str(row[6])
        if not name:
            skipped_other += 1
            continue

        # Parse services_json from string
        try:
            services = json.loads(row[16]) if row[16] else []
        except:
            services = []

        # Map shade_front/shade_rear to first valid value if shade is null
        for svc in services:
            if not svc.get('shade'):
                svc['shade'] = svc.get('shade_front') or svc.get('shade_rear')
            # add label for compatibility
            sk = svc.get('service_key', '')
            label_map = {
                'FULL_SIDES': 'Full (Sides & Rear)',
                'TWO_FRONT_DOORS': '2 Front Doors Only',
                'FULL_WS': 'Full Windshield',
                'SUN_STRIP': 'Sun Strip',
                'SUNROOF_SINGLE': 'Single Sunroof',
                'SUNROOF_PANO': 'Full Panoramic Sunroof',
                'REMOVAL_FULL': 'Removal',
            }
            svc['label'] = label_map.get(sk, sk)
            svc['price'] = svc.get('base_price', 0)
            svc['filmName'] = svc.get('film_type')

        # Map raw booking_source values to allowed enum
        raw_src = (safe_str(row[2]) or '').lower()
        if raw_src in ('nodep', 'shopify', 'online'):
            booking_src = 'online'
        elif raw_src in ('phone',):
            booking_src = 'phone'
        elif raw_src in ('gc', 'gift_certificate'):
            booking_src = 'gc'
        elif raw_src in ('sameday',):
            booking_src = 'sameday'
        else:
            booking_src = 'internal'

        # Map appointment_type
        raw_appt = (safe_str(row[4]) or '').lower()
        if raw_appt in ('dropoff', 'waiting', 'headsup_30', 'headsup_60'):
            appt_type = raw_appt
        else:
            appt_type = 'dropoff'

        booking = {
            'shop_id': SHOP_ID,
            'booking_id': booking_id,
            'booking_source': booking_src,
            'appointment_type': appt_type,
            'service_type': 'tint',
            'emoji_marker': safe_str(row[5]),
            'customer_name': name,
            'customer_email': normalize_email(row[7]),
            'customer_phone': normalize_phone(row[8]),
            'vehicle_year': int(row[9]) if row[9] else None,
            'vehicle_make': safe_str(row[10]),
            'vehicle_model': safe_str(row[11]),
            'class_keys': safe_str(row[12]),
            'appointment_date': d.date().isoformat(),
            'appointment_time': to_time_str(row[14]),
            'duration_minutes': int(row[15]) if row[15] else None,
            'services_json': services,
            'subtotal': to_num(row[18]),  # ORIGINAL booking total - frozen
            'discount_code': safe_str(row[19]),
            'discount_type': safe_str(row[20]) if safe_str(row[20]) in ('dollar', 'percent') else None,
            'discount_percent': to_num(row[21]),
            'discount_amount': to_num(row[22]),
            'deposit_paid': to_num(row[23]),
            'balance_due': to_num(row[24]),
            'calendar_event_id': safe_str(row[25]),
            # All past appointments are completed (Joe's instruction)
            'status': 'completed',
            'additional_interests': safe_str(row[28]),
            'notes': safe_str(row[30]),
            'window_status': safe_str(row[31]),
            'module': 'auto_tint',
            'import_source': 'historic_import',
        }
        rows_to_insert.append(booking)

    print(f"  Prepared {len(rows_to_insert)} rows ({skipped_future} future, {skipped_other} other)")
    inserted = insert_batch('auto_bookings', rows_to_insert, batch_size=200)
    print(f"  Inserted {inserted} bookings")
    return inserted


# ============================================================================
# IMPORT 2026FWT → auto_bookings + documents + line items + payments
# Match BookingLog rows to 2026FWT by Google Calendar event_id (the true link).
# When matched: use BookingLog.subtotal as starting_total (frozen original).
# When unmatched: fall back to 2026FWT.StartingTotal column (manually filled).
# ============================================================================
def import_2026fwt():
    print("\n=== Importing 2026FWT (final invoices + bookings) ===")
    wb = openpyxl.load_workbook(FINANCIAL_XLSX, data_only=True)
    ws = wb['2026FWT']

    # Build BookingLog lookup by Google Calendar event_id
    print("  Loading BookingLog for event_id matching...")
    wb_bl = openpyxl.load_workbook(BOOKINGLOG_XLSX, data_only=True)
    ws_bl = wb_bl['BookingLog']
    booking_log_by_event = {}
    for br in range(2, ws_bl.max_row + 1):
        bl_row = [c.value for c in ws_bl[br]]
        bl_event = bl_row[25]  # calendar_event_id
        bl_date = bl_row[13]   # appointment_date
        if bl_event and bl_date and hasattr(bl_date, 'year') and bl_date.year == 2026:
            try:
                services = json.loads(bl_row[16]) if bl_row[16] else []
            except:
                services = []
            # Normalize service objects with friendly fields
            for svc in services:
                if not svc.get('shade'):
                    svc['shade'] = svc.get('shade_front') or svc.get('shade_rear')
                sk = svc.get('service_key', '')
                label_map = {
                    'FULL_SIDES': 'Full (Sides & Rear)',
                    'TWO_FRONT_DOORS': '2 Front Doors Only',
                    'FULL_WS': 'Full Windshield',
                    'SUN_STRIP': 'Sun Strip',
                    'SUNROOF_SINGLE': 'Single Sunroof',
                    'SUNROOF_PANO': 'Full Panoramic Sunroof',
                    'REMOVAL_FULL': 'Removal',
                }
                svc['label'] = label_map.get(sk, sk)
                svc['price'] = svc.get('base_price', 0)
                svc['filmName'] = svc.get('film_type')
            booking_log_by_event[str(bl_event).strip()] = {
                'booking_id': bl_row[0],
                'name': bl_row[6],
                'phone': bl_row[8],
                'email': bl_row[7],
                'vehicle_year': bl_row[9],
                'vehicle_make': bl_row[10],
                'vehicle_model': bl_row[11],
                'class_keys': bl_row[12],
                'date': bl_date,
                'time': bl_row[14],
                'duration': bl_row[15],
                'subtotal': float(bl_row[18]) if bl_row[18] else 0,
                'services_json': services,
                'discount_code': bl_row[19],
                'discount_amount': float(bl_row[22]) if bl_row[22] else 0,
                'deposit_paid': float(bl_row[23]) if bl_row[23] else 0,
            }
    print(f"  Loaded {len(booking_log_by_event)} BookingLog entries with event_id")

    seen_doc_numbers = set()
    seen_booking_ids = set()
    booking_buffer = []  # auto_bookings records to insert
    doc_buffer = []  # documents (linked to bookings by row index after insert)
    matched_by_event = 0
    fallback_starting_total = 0
    no_starting_data = 0

    for r in range(2, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        d = row[0]
        if not d or not hasattr(d, 'year') or d.year != 2026:
            continue

        invoice_num = safe_str(row[29])
        if not invoice_num:
            continue

        # Look up matching BookingLog row by Google Calendar event_id
        event_id = safe_str(row[30])
        bl_match = booking_log_by_event.get(event_id) if event_id else None

        # Parse the row
        full_val = safe_str(row[6])
        fd2_val = safe_str(row[7])
        ws_val = safe_str(row[8])
        ts_val = safe_str(row[9])
        sr_val = safe_str(row[10])
        rem_val = safe_str(row[11])
        sub_pre = to_num(row[12])
        discount = to_num(row[13])
        discount_note = safe_str(row[14])
        nfw = to_num(row[15])
        tip = to_num(row[16])
        total = to_num(row[17])  # final post-discount, includes NFW + tip
        deposit = to_num(row[18])
        balance = to_num(row[19])  # what was actually due (post deposit/cash discount)
        payment_method = safe_str(row[20])
        processor = safe_str(row[21])
        starting_total_sheet = to_num(row[27])
        invoice_id_str = safe_str(row[29])

        # Determine the FINAL service subtotal (pre-discount, pre-warranty, pre-tip)
        # If sub_pre is filled, use it. Otherwise back-calculate from total.
        if sub_pre > 0:
            final_subtotal = sub_pre
        else:
            # No discount was applied -- total = services + nfw + tip
            final_subtotal = total - nfw - tip

        # Determine the original booking total (frozen-in-time value):
        # 1. If matched to BookingLog by event_id, use that subtotal (true source)
        # 2. Else if 2026FWT.StartingTotal is filled, use that (Joe's manual entry)
        # 3. Else use final_subtotal (no upsell trackable)
        if bl_match and bl_match['subtotal'] > 0:
            original_total = bl_match['subtotal']
            matched_by_event += 1
        elif starting_total_sheet > 0:
            original_total = starting_total_sheet
            fallback_starting_total += 1
        else:
            original_total = final_subtotal
            no_starting_data += 1

        # Calculate upsell: max(0, final services pre-discount - original)
        upsell = max(0, final_subtotal - original_total)

        # Use total for documents.subtotal? No -- documents.subtotal should be services pre-discount
        # so the cash discount math works correctly
        subtotal = final_subtotal

        # Build line items
        items = []
        sort_idx = 0
        if full_val:
            items.append({'description': f'Full (Sides & Rear) {full_val}', 'service_key': 'FULL_SIDES', 'sort': sort_idx})
            sort_idx += 1
        if fd2_val:
            items.append({'description': f'2 Front Doors {fd2_val}', 'service_key': 'TWO_FRONT_DOORS', 'sort': sort_idx})
            sort_idx += 1
        if ws_val:
            items.append({'description': f'Windshield {ws_val}', 'service_key': 'FULL_WS', 'sort': sort_idx})
            sort_idx += 1
        if ts_val:
            items.append({'description': f'Sun Strip {ts_val}', 'service_key': 'SUN_STRIP', 'sort': sort_idx})
            sort_idx += 1
        if sr_val:
            items.append({'description': f'Sunroof {sr_val}', 'service_key': 'SUNROOF_SINGLE', 'sort': sort_idx})
            sort_idx += 1
        if rem_val:
            items.append({'description': f'Removal', 'service_key': 'REMOVAL_FULL', 'sort': sort_idx})
            sort_idx += 1

        # If no items parsed, create one generic line for the subtotal
        if not items and subtotal > 0:
            items.append({'description': 'Service', 'service_key': 'OTHER', 'sort': 0})

        # Distribute the subtotal across items (best-effort)
        if items:
            per_item = subtotal / len(items)
            for it in items:
                it['line_total'] = round(per_item, 2)
                it['unit_price'] = round(per_item, 2)

        # Build a unique doc number (handle duplicates from spreadsheet)
        base_doc_num = f"INV-{invoice_id_str}"
        doc_num = base_doc_num
        suffix = 1
        while doc_num in seen_doc_numbers:
            suffix += 1
            doc_num = f"{base_doc_num}-{suffix}"
        seen_doc_numbers.add(doc_num)

        # Vehicle parsing: prefer BookingLog match, else parse from 2026FWT.Vehicle column
        if bl_match and bl_match.get('vehicle_make'):
            doc_vehicle_year = int(bl_match['vehicle_year']) if bl_match.get('vehicle_year') else None
            doc_vehicle_make = safe_str(bl_match['vehicle_make'])
            doc_vehicle_model = safe_str(bl_match['vehicle_model'])
        else:
            doc_vehicle_year, doc_vehicle_make, doc_vehicle_model = parse_vehicle(safe_str(row[4]))

        # All historic appointments are paid in full (we wouldn't have them in the financial sheet otherwise)
        document = {
            'shop_id': SHOP_ID,
            'doc_number': doc_num,
            'doc_type': 'invoice',
            'checkout_type': 'counter',
            'status': 'paid',
            'customer_name': safe_str(row[1]),
            'customer_phone': normalize_phone(row[2]),
            'customer_email': normalize_email(row[3]),
            'vehicle_year': doc_vehicle_year,
            'vehicle_make': doc_vehicle_make,
            'vehicle_model': doc_vehicle_model,
            'subtotal': subtotal,
            'starting_total': original_total,
            'upsell_amount': upsell,
            'discount_amount': discount,
            'deposit_paid': deposit,
            'balance_due': 0,
            'total_paid': total,
            'tip_amount': tip,
            'payment_method': (payment_method or 'cash').lower().replace(' ', '_'),
            'payment_processor': (processor or 'manual').lower(),
            'discount_note': discount_note,
            'tax_rate': 0,
            'tax_amount': 0,
            'paid_at': d.isoformat(),
            'created_at': d.isoformat(),
            'import_source': 'historic_import',
        }

        # Build the auto_bookings record (the original booking)
        # Use unique booking_id with row suffix to avoid spreadsheet duplicates
        base_booking_id = invoice_num
        booking_id_unique = base_booking_id
        suffix = 1
        while booking_id_unique in seen_booking_ids:
            suffix += 1
            booking_id_unique = f"{base_booking_id}-{suffix}"
        seen_booking_ids.add(booking_id_unique)

        # Build the booking record with consistent key set across all rows
        booking_record = {
            'shop_id': SHOP_ID,
            'booking_id': booking_id_unique,
            'booking_source': 'online',
            'appointment_type': 'dropoff',
            'service_type': 'tint',
            'customer_name': (bl_match['name'] if bl_match else None) or safe_str(row[1]),
            'customer_phone': normalize_phone((bl_match['phone'] if bl_match else None) or row[2]),
            'customer_email': normalize_email((bl_match['email'] if bl_match else None) or row[3]),
            'vehicle_year': int(bl_match['vehicle_year']) if bl_match and bl_match.get('vehicle_year') else None,
            'vehicle_make': safe_str(bl_match['vehicle_make']) if bl_match else None,
            'vehicle_model': safe_str(bl_match['vehicle_model']) if bl_match else None,
            'class_keys': safe_str(bl_match['class_keys']) if bl_match else None,
            'appointment_date': (bl_match['date'].date() if bl_match else d.date()).isoformat(),
            'appointment_time': to_time_str(bl_match['time']) if bl_match else None,
            'duration_minutes': int(bl_match['duration']) if bl_match and bl_match.get('duration') else None,
            'services_json': bl_match['services_json'] if bl_match else [],
            'subtotal': original_total,
            'discount_code': safe_str(bl_match['discount_code']) if bl_match else None,
            'discount_amount': bl_match['discount_amount'] if bl_match else 0,
            'deposit_paid': bl_match['deposit_paid'] if bl_match else 0,
            'balance_due': 0,
            'total_paid': total,
            'payment_method': (payment_method or 'cash').lower().replace(' ', '_'),
            'calendar_event_id': event_id,
            'status': 'completed',
            'module': 'auto_tint',
            'import_source': 'historic_import',
            'notes': discount_note,
        }

        booking_buffer.append(booking_record)

        doc_buffer.append({
            'doc': document,
            'booking_index': len(booking_buffer) - 1,  # Will resolve to ID after batch insert
            'items': items,
            'payment': {
                'method': payment_method,
                'processor': processor,
                'amount': total,
                'tip': tip,
                'date': d.isoformat(),
            } if total > 0 else None
        })

    print(f"  Prepared {len(booking_buffer)} bookings + {len(doc_buffer)} documents")
    print(f"    Original total source: {matched_by_event} matched via event_id, {fallback_starting_total} from sheet StartingTotal, {no_starting_data} no original (upsell=0)")

    # Step 1: Insert all auto_bookings first (in batches), keep order
    print("  Inserting bookings...")
    inserted_booking_ids = []
    batch_size = 100
    for i in range(0, len(booking_buffer), batch_size):
        batch = booking_buffer[i:i+batch_size]
        resp = supabase_request("POST", "auto_bookings", body=batch)
        for ins in resp:
            inserted_booking_ids.append(ins['id'])
        if i % 200 == 0 and i > 0:
            print(f"    {len(inserted_booking_ids)}/{len(booking_buffer)}")

    # Step 2: Insert documents in batches
    print("  Inserting documents...")
    inserted_docs = []
    for i in range(0, len(doc_buffer), batch_size):
        batch = doc_buffer[i:i+batch_size]
        docs_only = [b['doc'] for b in batch]
        resp = supabase_request("POST", "documents", body=docs_only)
        for j, inserted in enumerate(resp):
            inserted_docs.append({
                'doc_id': inserted['id'],
                'booking_id': inserted_booking_ids[batch[j]['booking_index']],
                'items': batch[j]['items'],
                'payment': batch[j]['payment'],
            })
        if i % 200 == 0 and i > 0:
            print(f"    {len(inserted_docs)}/{len(doc_buffer)}")

    # Build line items and payments now that we have doc_ids
    line_items = []
    payments = []
    for entry in inserted_docs:
        for it in entry['items']:
            line_items.append({
                'document_id': entry['doc_id'],
                'module': 'auto_tint',
                'description': it['description'],
                'quantity': 1,
                'unit_price': it['line_total'],
                'line_total': it['line_total'],
                'sort_order': it['sort'],
                'custom_fields': {'serviceKey': it['service_key']},
            })
        if entry['payment']:
            p = entry['payment']
            payments.append({
                'document_id': entry['doc_id'],
                'shop_id': SHOP_ID,
                'amount': p['amount'],
                'payment_method': (p['method'] or 'cash').lower().replace(' ', '_'),
                'processor': (p['processor'] or 'manual').lower(),
                'status': 'confirmed',
                'created_at': p['date'],
            })

    print(f"  Inserting {len(line_items)} line items...")
    insert_batch('document_line_items', line_items, batch_size=200)

    print(f"  Inserting {len(payments)} payment records...")
    insert_batch('document_payments', payments, batch_size=200)

    # Update auto_bookings.document_id
    print("  Linking bookings to documents...")
    for entry in inserted_docs:
        supabase_request("PATCH", "auto_bookings",
            body={'document_id': entry['doc_id']},
            params={'id': f"eq.{entry['booking_id']}"})

    return len(inserted_docs)


# ============================================================================
# MAIN
# ============================================================================
def main():
    load_env()
    print(f"Connected: {SUPABASE_URL}")

    if "--year" in sys.argv:
        year = int(sys.argv[sys.argv.index("--year") + 1])
        if year in (2024, 2025):
            import_yearly_sheet(year)
        return

    if "--bookinglog" in sys.argv:
        import_bookinglog_2026()
        return

    if "--2026fwt" in sys.argv:
        import_2026fwt()
        return

    if "--all" in sys.argv:
        import_yearly_sheet(2024)
        import_yearly_sheet(2025)
        import_bookinglog_2026()
        import_2026fwt()
        return

    print("Usage: python3 import_historic_data.py [--year 2024|2025 | --bookinglog | --2026fwt | --all]")

if __name__ == '__main__':
    main()
